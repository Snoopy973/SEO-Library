import streamlit as st
import pandas as pd
import json
import re
import ssl
import urllib.request
import urllib.error
from collections import Counter, defaultdict
from io import BytesIO
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import PieChart, Reference
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

SSL_CTX = ssl.create_default_context()
SSL_CTX.check_hostname = False
SSL_CTX.verify_mode = ssl.CERT_NONE

st.title("🛒 Scraper Produits — Shopify & WooCommerce")
st.markdown("Scrapez un site e-commerce ou uploadez un JSON produit pour analyser le catalogue.")

# ─────────────────────────────────────────────
# FONCTIONS SCRAPING
# ─────────────────────────────────────────────

def fetch_shopify_products(store_domain, progress_bar=None):
    all_products = []
    page = 1
    store_domain = store_domain.replace("https://", "").replace("http://", "").strip("/")
    if not store_domain.startswith("www."):
        store_domain = "www." + store_domain
    base_url = f"https://{store_domain}/products.json?limit=250"

    while True:
        url = f"{base_url}&page={page}"
        try:
            req = urllib.request.Request(url, headers={
                "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7)"
            })
            with urllib.request.urlopen(req, timeout=30, context=SSL_CTX) as response:
                data = json.loads(response.read().decode())
            products = data.get("products", [])
            if not products:
                break
            all_products.extend(products)
            if progress_bar:
                progress_bar.progress(min(page * 0.15, 0.95), f"Page {page} — {len(all_products)} produits...")
            page += 1
        except urllib.error.HTTPError as e:
            if e.code == 429:
                import time; time.sleep(2)
                continue
            else:
                st.error(f"Erreur HTTP {e.code}: {e.reason}")
                break
        except Exception as e:
            st.error(f"Erreur: {e}")
            break

    if progress_bar:
        progress_bar.progress(1.0, f"Terminé — {len(all_products)} produits")
    return all_products


def fetch_woo_products(store_domain, progress_bar=None):
    all_products = []
    page = 1
    store_domain = store_domain.replace("https://", "").replace("http://", "").strip("/")
    if not store_domain.startswith("www."):
        store_domain = "www." + store_domain
    base_url = f"https://{store_domain}/wp-json/wc/store/products?per_page=100"

    while True:
        url = f"{base_url}&page={page}"
        try:
            req = urllib.request.Request(url, headers={
                "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7)"
            })
            with urllib.request.urlopen(req, timeout=30, context=SSL_CTX) as response:
                data = json.loads(response.read().decode())
            if not data:
                break
            products = data if isinstance(data, list) else data.get("products", data.get("items", []))
            if not products:
                break
            all_products.extend(products)
            if progress_bar:
                progress_bar.progress(min(page * 0.15, 0.95), f"Page {page} — {len(all_products)} produits...")
            page += 1
        except urllib.error.HTTPError as e:
            if e.code == 400 or e.code == 404:
                break
            elif e.code == 429:
                import time; time.sleep(2)
                continue
            else:
                st.error(f"Erreur HTTP {e.code}: {e.reason}")
                break
        except Exception as e:
            st.error(f"Erreur: {e}")
            break

    if progress_bar:
        progress_bar.progress(1.0, f"Terminé — {len(all_products)} produits")
    return all_products


# ─────────────────────────────────────────────
# FONCTIONS EXTRACTION
# ─────────────────────────────────────────────

def extract_tag_values(tags, prefix):
    if isinstance(tags, list):
        return [t.split(":", 1)[1].strip().capitalize()
                for t in tags if isinstance(t, str) and t.lower().startswith(prefix + ":")]
    return []


def extract_materials_from_description(html_body):
    if not html_body:
        return []
    text = re.sub(r'<[^>]+>', ' ', html_body).replace('&nbsp;', ' ')
    materials = []
    for match in re.findall(r'(\d+)%\s*([\wéèêë]+)', text.lower()):
        mat = match[1].strip().capitalize()
        if mat not in materials:
            materials.append(mat)
    return materials


def extract_composition(html_body):
    if not html_body:
        return ""
    text = re.sub(r'<[^>]+>', ' ', html_body)
    matches = re.findall(r'\d+%\s*[\wéèêë]+', text.lower())
    return ", ".join(matches) if matches else ""


def detect_cms(products):
    """Détecte si c'est du Shopify ou WooCommerce"""
    if not products:
        return "unknown"
    sample = products[0]
    if "variants" in sample and "vendor" in sample:
        return "shopify"
    if "prices" in sample or "price_html" in sample:
        return "woocommerce"
    return "unknown"


def parse_shopify_product(p, store_domain):
    tags = p.get("tags", [])
    body = p.get("body_html", "")
    product_type = p.get("product_type", "Non défini")
    materials = extract_tag_values(tags, "matiere") or extract_tag_values(tags, "matière")
    if not materials:
        materials = extract_materials_from_description(body)
    colors = extract_tag_values(tags, "couleur")
    coupes = extract_tag_values(tags, "coupe")
    formes = extract_tag_values(tags, "forme")
    collections = [t.split(":", 1)[1].strip() for t in tags
                   if isinstance(t, str) and t.lower().startswith("collection:")]

    variants = p.get("variants", [])
    price = float(variants[0].get("price", 0)) if variants else None
    compare_price = None
    if variants and variants[0].get("compare_at_price"):
        compare_price = float(variants[0]["compare_at_price"])

    return {
        "title": p.get("title", ""),
        "type": product_type,
        "materials": materials,
        "materials_str": ", ".join(materials) if materials else "Non renseigné",
        "composition": extract_composition(body),
        "colors": colors,
        "colors_str": ", ".join(colors) if colors else "Non renseigné",
        "coupes": coupes,
        "coupes_str": ", ".join(coupes) if coupes else "",
        "formes": formes,
        "formes_str": ", ".join(formes) if formes else "",
        "collections": collections,
        "price": price,
        "compare_price": compare_price,
        "url": f"https://{store_domain}/products/{p.get('handle', '')}",
        "tags": tags,
    }


def parse_woo_product(p, store_domain):
    name = p.get("name", p.get("title", ""))
    body = p.get("description", p.get("short_description", ""))
    categories = [c.get("name", "") for c in p.get("categories", [])]
    product_type = categories[0] if categories else "Non défini"
    tags = [t.get("name", "") for t in p.get("tags", [])]
    materials = extract_materials_from_description(body)

    price = None
    prices = p.get("prices", {})
    if prices:
        price_str = prices.get("price", prices.get("regular_price", "0"))
        try:
            price = float(price_str) / 100 if len(str(price_str)) > 4 else float(price_str)
        except (ValueError, TypeError):
            pass
    elif p.get("price"):
        try:
            price = float(p["price"])
        except (ValueError, TypeError):
            pass

    permalink = p.get("permalink", p.get("slug", ""))
    if permalink and not permalink.startswith("http"):
        permalink = f"https://{store_domain}/product/{permalink}"

    return {
        "title": name,
        "type": product_type,
        "materials": materials,
        "materials_str": ", ".join(materials) if materials else "Non renseigné",
        "composition": extract_composition(body),
        "colors": [],
        "colors_str": "Non renseigné",
        "coupes": [],
        "coupes_str": "",
        "formes": [],
        "formes_str": "",
        "collections": categories,
        "price": price,
        "compare_price": None,
        "url": permalink,
        "tags": tags,
    }


# ─────────────────────────────────────────────
# ANALYSE COMPLÈTE
# ─────────────────────────────────────────────

def analyze_parsed_products(parsed_products):
    results = {
        "products": parsed_products,
        "materials_count": Counter(),
        "type_count": Counter(),
        "color_count": Counter(),
        "coupe_count": Counter(),
        "forme_count": Counter(),
        "collection_count": Counter(),
        "material_by_type": defaultdict(Counter),
        "coupe_by_type": defaultdict(Counter),
        "price_by_material": defaultdict(list),
        "combos_type_mat": Counter(),
        "combos_type_col": Counter(),
        "combos_type_coupe": Counter(),
        "taxonomy": defaultdict(set),
        "total": len(parsed_products),
    }

    for p in parsed_products:
        ptype = p["type"]
        results["type_count"][ptype] += 1
        results["taxonomy"]["Type de produit"].add(ptype)

        for mat in p["materials"]:
            results["materials_count"][mat] += 1
            results["material_by_type"][mat][ptype] += 1
            results["taxonomy"]["Matières"].add(mat)
            if p["price"]:
                results["price_by_material"][mat].append(p["price"])
            results["combos_type_mat"][f"{ptype.lower()} {mat.lower()}"] += 1

        for col in p["colors"]:
            results["color_count"][col] += 1
            results["taxonomy"]["Couleurs"].add(col)
            results["combos_type_col"][f"{ptype.lower()} {col.lower()}"] += 1

        for coupe in p["coupes"]:
            results["coupe_count"][coupe] += 1
            results["coupe_by_type"][coupe][ptype] += 1
            results["taxonomy"]["Coupes"].add(coupe)
            results["combos_type_coupe"][f"{ptype.lower()} {coupe.lower()}"] += 1

        for forme in p["formes"]:
            results["forme_count"][forme] += 1
            results["taxonomy"]["Formes"].add(forme)

        for coll in p["collections"]:
            results["collection_count"][coll] += 1
            results["taxonomy"]["Collections"].add(coll)

    return results


# ─────────────────────────────────────────────
# UI
# ─────────────────────────────────────────────

tab_scrape, tab_upload = st.tabs(["🌐 Scraper un site", "📁 Uploader un JSON"])

# ── TAB 1 : Scraper ──
with tab_scrape:
    col1, col2 = st.columns([3, 1])
    with col1:
        store_url = st.text_input("URL du site e-commerce",
                                   placeholder="www.balibaris.com ou www.mon-site-woo.com",
                                   help="Fonctionne avec Shopify et WooCommerce")
    with col2:
        cms_choice = st.selectbox("CMS", ["Auto-détecter", "Shopify", "WooCommerce"])

    if st.button("🚀 Lancer le scraping", type="primary", use_container_width=True):
        if not store_url:
            st.warning("Entre une URL")
        else:
            progress = st.progress(0, "Démarrage...")
            domain = store_url.replace("https://", "").replace("http://", "").strip("/")

            if cms_choice == "WooCommerce":
                raw_products = fetch_woo_products(domain, progress)
                cms = "woocommerce"
            elif cms_choice == "Shopify":
                raw_products = fetch_shopify_products(domain, progress)
                cms = "shopify"
            else:
                raw_products = fetch_shopify_products(domain, progress)
                cms = "shopify"
                if not raw_products:
                    st.info("Pas un Shopify, tentative WooCommerce...")
                    raw_products = fetch_woo_products(domain, progress)
                    cms = "woocommerce"

            if raw_products:
                st.success(f"✅ {len(raw_products)} produits récupérés ({cms})")
                store_clean = domain.replace("www.", "")
                parser = parse_shopify_product if cms == "shopify" else parse_woo_product
                parsed = [parser(p, domain) for p in raw_products]
                results = analyze_parsed_products(parsed)
                st.session_state["product_results"] = results
                st.session_state["product_store"] = store_clean
                st.session_state["product_raw_json"] = raw_products
            else:
                st.error("❌ Aucun produit trouvé. Vérifiez l'URL et le CMS.")

# ── TAB 2 : Upload JSON ──
with tab_upload:
    uploaded = st.file_uploader("Upload un fichier JSON produits", type=["json"],
                                 help="Export Shopify /products.json ou WooCommerce /wp-json/wc/store/products")
    if uploaded:
        try:
            data = json.loads(uploaded.read().decode())
            if isinstance(data, dict) and "products" in data:
                raw_products = data["products"]
            elif isinstance(data, list):
                raw_products = data
            else:
                raw_products = []

            if raw_products:
                cms = detect_cms(raw_products)
                st.success(f"✅ {len(raw_products)} produits chargés (détecté: {cms})")

                store_name = st.text_input("Nom du site", value=uploaded.name.replace(".json", ""))
                domain = store_name.replace("www.", "").replace(".com", "").replace(".fr", "")
                parser = parse_shopify_product if cms == "shopify" else parse_woo_product
                parsed = [parser(p, store_name) for p in raw_products]
                results = analyze_parsed_products(parsed)
                st.session_state["product_results"] = results
                st.session_state["product_store"] = domain
                st.session_state["product_raw_json"] = raw_products
            else:
                st.error("Aucun produit trouvé dans ce fichier")
        except Exception as e:
            st.error(f"Erreur de lecture : {e}")


# ─────────────────────────────────────────────
# AFFICHAGE DES RÉSULTATS
# ─────────────────────────────────────────────

if "product_results" in st.session_state:
    results = st.session_state["product_results"]
    store = st.session_state.get("product_store", "site")
    total = results["total"]

    st.divider()
    st.subheader(f"📊 Analyse de {store} — {total} produits")

    # KPIs
    c1, c2, c3, c4, c5 = st.columns(5)
    c1.metric("Produits", total)
    c2.metric("Types", len(results["type_count"]))
    c3.metric("Matières", len(results["materials_count"]))
    c4.metric("Couleurs", len(results["color_count"]))
    c5.metric("Coupes", len(results["coupe_count"]))

    # Onglets visuels
    t1, t2, t3, t4, t5, t6, t7 = st.tabs([
        "🧵 Matières", "📦 Types", "🎨 Couleurs", "📐 Coupes",
        "🔀 Croisements", "🗂 Taxonomie", "📋 Tous"
    ])

    with t1:
        if results["materials_count"]:
            df_mat = pd.DataFrame(results["materials_count"].most_common(),
                                   columns=["Matière", "Nb produits"])
            df_mat["% du catalogue"] = (df_mat["Nb produits"] / total * 100).round(1)

            # Prix par matière
            prix_data = []
            for mat in df_mat["Matière"]:
                prices = results["price_by_material"].get(mat, [])
                prix_data.append({
                    "Prix moyen": round(sum(prices) / len(prices), 2) if prices else None,
                    "Prix min": round(min(prices), 2) if prices else None,
                    "Prix max": round(max(prices), 2) if prices else None,
                })
            df_prix = pd.DataFrame(prix_data)
            df_mat = pd.concat([df_mat, df_prix], axis=1)

            col1, col2 = st.columns([2, 1])
            with col1:
                st.dataframe(df_mat, use_container_width=True, hide_index=True)
            with col2:
                import plotly.express as px
                fig = px.pie(df_mat.head(15), values="Nb produits", names="Matière",
                            title="Top 15 matières")
                fig.update_layout(height=400)
                st.plotly_chart(fig, use_container_width=True)

    with t2:
        if results["type_count"]:
            df_type = pd.DataFrame(results["type_count"].most_common(),
                                    columns=["Type", "Nb produits"])
            df_type["% du catalogue"] = (df_type["Nb produits"] / total * 100).round(1)
            col1, col2 = st.columns([2, 1])
            with col1:
                st.dataframe(df_type, use_container_width=True, hide_index=True)
            with col2:
                import plotly.express as px
                fig = px.bar(df_type.head(15), x="Type", y="Nb produits", title="Par type")
                fig.update_layout(height=400)
                st.plotly_chart(fig, use_container_width=True)

    with t3:
        if results["color_count"]:
            df_col = pd.DataFrame(results["color_count"].most_common(),
                                   columns=["Couleur", "Nb produits"])
            df_col["% du catalogue"] = (df_col["Nb produits"] / total * 100).round(1)
            col1, col2 = st.columns([2, 1])
            with col1:
                st.dataframe(df_col, use_container_width=True, hide_index=True)
            with col2:
                import plotly.express as px
                fig = px.bar(df_col.head(20), x="Couleur", y="Nb produits", title="Top 20 couleurs")
                fig.update_layout(height=400)
                st.plotly_chart(fig, use_container_width=True)

    with t4:
        if results["coupe_count"]:
            df_coupe = pd.DataFrame(results["coupe_count"].most_common(),
                                     columns=["Coupe", "Nb produits"])
            df_coupe["% du catalogue"] = (df_coupe["Nb produits"] / total * 100).round(1)
            col1, col2 = st.columns([2, 1])
            with col1:
                st.dataframe(df_coupe, use_container_width=True, hide_index=True)
            with col2:
                import plotly.express as px
                fig = px.pie(df_coupe, values="Nb produits", names="Coupe", title="Répartition coupes")
                fig.update_layout(height=400)
                st.plotly_chart(fig, use_container_width=True)

            # Croisement coupe x type
            if results["coupe_by_type"]:
                st.markdown("#### Croisement Coupe x Type")
                all_types = sorted(set(t for c in results["coupe_by_type"].values() for t in c))
                cross_data = []
                for coupe, tc in sorted(results["coupe_by_type"].items()):
                    row = {"Coupe": coupe}
                    for t in all_types:
                        row[t] = tc.get(t, 0)
                    cross_data.append(row)
                st.dataframe(pd.DataFrame(cross_data), use_container_width=True, hide_index=True)
        else:
            st.info("Aucune donnée de coupe trouvée dans les tags produits")

    with t5:
        st.markdown("#### Matière x Type de produit")
        if results["material_by_type"]:
            all_types = sorted(set(t for c in results["material_by_type"].values() for t in c))
            cross_data = []
            for mat, tc in sorted(results["material_by_type"].items()):
                row = {"Matière": mat}
                for t in all_types:
                    row[t] = tc.get(t, 0)
                cross_data.append(row)
            df_cross = pd.DataFrame(cross_data)
            st.dataframe(df_cross, use_container_width=True, hide_index=True)

            # Heatmap
            import plotly.express as px
            df_heat = df_cross.set_index("Matière")
            fig = px.imshow(df_heat.head(20), text_auto=True, aspect="auto",
                           title="Heatmap Matière x Type (top 20 matières)")
            fig.update_layout(height=600)
            st.plotly_chart(fig, use_container_width=True)

    with t6:
        st.markdown("#### Taxonomie — Toutes les valeurs uniques par attribut")
        taxonomy = results["taxonomy"]
        if taxonomy:
            max_len = max(len(v) for v in taxonomy.values())
            tax_data = {}
            for col_name, values in sorted(taxonomy.items()):
                sorted_vals = sorted(values)
                tax_data[f"{col_name} ({len(sorted_vals)})"] = sorted_vals + [""] * (max_len - len(sorted_vals))
            st.dataframe(pd.DataFrame(tax_data), use_container_width=True, hide_index=True)

    with t7:
        st.markdown("#### Liste complète des produits")
        df_all = pd.DataFrame([{
            "Produit": p["title"],
            "Type": p["type"],
            "Matières": p["materials_str"],
            "Couleurs": p["colors_str"],
            "Coupes": p["coupes_str"],
            "Prix": p["price"],
            "URL": p["url"],
        } for p in results["products"]])
        st.dataframe(df_all, use_container_width=True, hide_index=True,
                      column_config={"URL": st.column_config.LinkColumn("URL")})

    # ── Export Excel ──
    st.divider()
    st.subheader("📥 Export")

    if st.button("💾 Générer le fichier Excel complet", type="primary"):
        # On réutilise la logique du script existant
        from io import BytesIO
        buffer = BytesIO()
        wb = Workbook()

        def style_h(ws, row, max_col):
            hf = PatternFill(start_color="1B2A4A", end_color="1B2A4A", fill_type="solid")
            hfont = Font(bold=True, color="FFFFFF", size=11)
            for c in range(1, max_col + 1):
                cell = ws.cell(row=row, column=c)
                cell.fill = hf; cell.font = hfont
                cell.alignment = Alignment(horizontal="center")

        # Onglet Matières
        ws1 = wb.active; ws1.title = "Par Matière"
        ws1.append(["Matière", "Nb produits", "% catalogue", "Prix moyen", "Prix min", "Prix max"])
        style_h(ws1, 1, 6)
        for mat, count in results["materials_count"].most_common():
            prices = results["price_by_material"].get(mat, [])
            ws1.append([mat, count, round(count/total*100, 1),
                        round(sum(prices)/len(prices), 2) if prices else "",
                        round(min(prices), 2) if prices else "",
                        round(max(prices), 2) if prices else ""])

        # Onglet Types
        ws2 = wb.create_sheet("Par Type")
        ws2.append(["Type", "Nb produits", "% catalogue"])
        style_h(ws2, 1, 3)
        for t, c in results["type_count"].most_common():
            ws2.append([t, c, round(c/total*100, 1)])

        # Onglet Couleurs
        ws3 = wb.create_sheet("Par Couleur")
        ws3.append(["Couleur", "Nb produits", "% catalogue"])
        style_h(ws3, 1, 3)
        for col, c in results["color_count"].most_common():
            ws3.append([col, c, round(c/total*100, 1)])

        # Onglet Coupes
        ws4 = wb.create_sheet("Par Coupe")
        ws4.append(["Coupe", "Nb produits", "% catalogue"])
        style_h(ws4, 1, 3)
        for cp, c in results["coupe_count"].most_common():
            ws4.append([cp, c, round(c/total*100, 1)])

        # Onglet Tous les produits
        ws5 = wb.create_sheet("Tous les produits")
        ws5.append(["Produit", "Type", "Matières", "Composition", "Couleurs", "Coupes", "Prix", "URL"])
        style_h(ws5, 1, 8)
        for p in results["products"]:
            ws5.append([p["title"], p["type"], p["materials_str"], p["composition"],
                        p["colors_str"], p["coupes_str"], p["price"], p["url"]])

        wb.save(buffer)
        ts = datetime.now().strftime("%Y%m%d_%H%M")
        fname = f"analyse_{store}_{ts}.xlsx"
        st.download_button("⬇️ Télécharger l'Excel", data=buffer.getvalue(),
                            file_name=fname, mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
