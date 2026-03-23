import streamlit as st
import pandas as pd
import json
import re
import ssl
import urllib.request
import urllib.error
from collections import Counter, defaultdict
from io import BytesIO, StringIO
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.chart import BarChart, PieChart, Reference
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

SSL_CTX = ssl.create_default_context()
SSL_CTX.check_hostname = False
SSL_CTX.verify_mode = ssl.CERT_NONE

st.title("📊 Analyse SEO Complète")
st.markdown("Scrape un site, charge tes exports Ahrefs, et obtiens l'analyse croisée en un seul endroit.")

# ═════════════════════════════════════════════
# FONCTIONS SCRAPING
# ═════════════════════════════════════════════

def fetch_shopify_products(store_domain, progress_bar=None):
    all_products = []
    page = 1
    store_domain = store_domain.replace("https://", "").replace("http://", "").strip("/")
    if not store_domain.startswith("www."):
        store_domain = "www." + store_domain
    base_url = f"https://{store_domain}/products.json?limit=250"
    while True:
        url = f"{base_url}&page={page}"
        try:
            req = urllib.request.Request(url, headers={
                "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7)"
            })
            with urllib.request.urlopen(req, timeout=30, context=SSL_CTX) as response:
                data = json.loads(response.read().decode())
            products = data.get("products", [])
            if not products:
                break
            all_products.extend(products)
            if progress_bar:
                progress_bar.progress(min(page * 0.15, 0.95), f"Page {page} — {len(all_products)} produits...")
            page += 1
        except urllib.error.HTTPError as e:
            if e.code == 429:
                import time; time.sleep(2)
                continue
            else:
                st.error(f"Erreur HTTP {e.code}: {e.reason}")
                break
        except Exception as e:
            st.error(f"Erreur: {e}")
            break
    if progress_bar:
        progress_bar.progress(1.0, f"Terminé — {len(all_products)} produits")
    return all_products


def fetch_woo_products(store_domain, progress_bar=None):
    all_products = []
    page = 1
    store_domain = store_domain.replace("https://", "").replace("http://", "").strip("/")
    if not store_domain.startswith("www."):
        store_domain = "www." + store_domain
    base_url = f"https://{store_domain}/wp-json/wc/store/products?per_page=100"
    while True:
        url = f"{base_url}&page={page}"
        try:
            req = urllib.request.Request(url, headers={
                "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7)"
            })
            with urllib.request.urlopen(req, timeout=30, context=SSL_CTX) as response:
                data = json.loads(response.read().decode())
            if not data:
                break
            products = data if isinstance(data, list) else data.get("products", data.get("items", []))
            if not products:
                break
            all_products.extend(products)
            if progress_bar:
                progress_bar.progress(min(page * 0.15, 0.95), f"Page {page} — {len(all_products)} produits...")
            page += 1
        except urllib.error.HTTPError as e:
            if e.code in (400, 404):
                break
            elif e.code == 429:
                import time; time.sleep(2)
                continue
            else:
                st.error(f"Erreur HTTP {e.code}: {e.reason}")
                break
        except Exception as e:
            st.error(f"Erreur: {e}")
            break
    if progress_bar:
        progress_bar.progress(1.0, f"Terminé — {len(all_products)} produits")
    return all_products


# ═════════════════════════════════════════════
# FONCTIONS EXTRACTION
# ═════════════════════════════════════════════

# All known tag prefixes for Shopify fashion sites
TAG_PREFIXES = {
    "matiere": "matières", "matière": "matières",
    "couleur": "couleurs",
    "coupe": "coupes",
    "forme": "formes",
    "motif": "motifs",
    "collection": "collections",
    "saison": "saisons",
    "genre": "genre",
    "guide": "guides_tailles",
    "categorie": "categories", "catégorie": "categories",
}


def extract_tag_values(tags, prefix):
    if isinstance(tags, list):
        return [t.split(":", 1)[1].strip().capitalize()
                for t in tags if isinstance(t, str) and t.lower().startswith(prefix + ":")]
    return []


def extract_all_tag_groups(tags):
    """Extract all tag prefix groups dynamically."""
    groups = defaultdict(list)
    if not isinstance(tags, list):
        return groups
    for t in tags:
        if not isinstance(t, str) or ":" not in t:
            continue
        prefix, value = t.split(":", 1)
        prefix = prefix.strip().lower()
        value = value.strip()
        if value:
            # Map to canonical group name
            group = TAG_PREFIXES.get(prefix, prefix)
            cap_val = value.capitalize() if not value[0].isupper() else value
            if cap_val not in groups[group]:
                groups[group].append(cap_val)
    return groups


def extract_materials_from_description(html_body):
    if not html_body:
        return []
    text = re.sub(r'<[^>]+>', ' ', html_body).replace('&nbsp;', ' ')
    materials = []
    for match in re.findall(r'(\d+)%\s*([\wéèêë]+)', text.lower()):
        mat = match[1].strip().capitalize()
        if mat not in materials:
            materials.append(mat)
    return materials


def extract_composition(html_body):
    if not html_body:
        return ""
    text = re.sub(r'<[^>]+>', ' ', html_body)
    matches = re.findall(r'\d+%\s*[\wéèêë]+', text.lower())
    return ", ".join(matches) if matches else ""


def parse_shopify_product(p, store_domain):
    tags = p.get("tags", [])
    body = p.get("body_html", "")
    product_type = p.get("product_type", "Non défini")

    # Extract all tag groups dynamically
    tag_groups = extract_all_tag_groups(tags)

    materials = tag_groups.get("matières", [])
    if not materials:
        materials = extract_materials_from_description(body)
    colors = tag_groups.get("couleurs", [])
    coupes = tag_groups.get("coupes", [])
    formes = tag_groups.get("formes", [])
    motifs = tag_groups.get("motifs", [])
    collections = tag_groups.get("collections", [])
    saisons = tag_groups.get("saisons", [])
    genre = tag_groups.get("genre", [])
    guides = tag_groups.get("guides_tailles", [])
    categories = tag_groups.get("categories", [])

    variants = p.get("variants", [])
    price = float(variants[0].get("price", 0)) if variants else None
    compare_price = None
    if variants and variants[0].get("compare_at_price"):
        compare_price = float(variants[0]["compare_at_price"])

    in_stock = any(v.get("available", False) for v in variants)
    available_variants = sum(1 for v in variants if v.get("available", False))
    total_variants = len(variants)

    return {
        "title": p.get("title", ""),
        "type": product_type,
        "materials": materials,
        "materials_str": ", ".join(materials) if materials else "Non renseigné",
        "composition": extract_composition(body),
        "colors": colors,
        "colors_str": ", ".join(colors) if colors else "Non renseigné",
        "coupes": coupes,
        "coupes_str": ", ".join(coupes) if coupes else "",
        "formes": formes,
        "formes_str": ", ".join(formes) if formes else "",
        "motifs": motifs,
        "motifs_str": ", ".join(motifs) if motifs else "",
        "collections": collections,
        "saisons": saisons,
        "genre": genre,
        "guides": guides,
        "categories": categories,
        "price": price,
        "compare_price": compare_price,
        "url": f"https://{store_domain}/products/{p.get('handle', '')}",
        "tags": tags,
        "in_stock": in_stock,
        "available_variants": available_variants,
        "total_variants": total_variants,
        "published_at": p.get("published_at"),
    }


def parse_woo_product(p, store_domain):
    name = p.get("name", p.get("title", ""))
    body = p.get("description", p.get("short_description", ""))
    categories = [c.get("name", "") for c in p.get("categories", [])]
    product_type = categories[0] if categories else "Non défini"
    tags = [t.get("name", "") for t in p.get("tags", [])]
    materials = extract_materials_from_description(body)
    price = None
    prices = p.get("prices", {})
    if prices:
        price_str = prices.get("price", prices.get("regular_price", "0"))
        try:
            price = float(price_str) / 100 if len(str(price_str)) > 4 else float(price_str)
        except (ValueError, TypeError):
            pass
    elif p.get("price"):
        try:
            price = float(p["price"])
        except (ValueError, TypeError):
            pass
    permalink = p.get("permalink", p.get("slug", ""))
    if permalink and not permalink.startswith("http"):
        permalink = f"https://{store_domain}/product/{permalink}"
    return {
        "title": name, "type": product_type,
        "materials": materials, "materials_str": ", ".join(materials) if materials else "Non renseigné",
        "composition": extract_composition(body),
        "colors": [], "colors_str": "Non renseigné",
        "coupes": [], "coupes_str": "", "formes": [], "formes_str": "",
        "motifs": [], "motifs_str": "",
        "collections": categories, "saisons": [], "genre": [], "guides": [], "categories": [],
        "price": price, "compare_price": None, "url": permalink, "tags": tags,
    }


def detect_cms(products):
    if not products:
        return "unknown"
    sample = products[0]
    if "variants" in sample and "vendor" in sample:
        return "shopify"
    if "prices" in sample or "price_html" in sample:
        return "woocommerce"
    return "unknown"


# ═════════════════════════════════════════════
# ANALYSE PRODUITS
# ═════════════════════════════════════════════

def analyze_parsed_products(parsed_products):
    results = {
        "products": parsed_products,
        "materials_count": Counter(),
        "type_count": Counter(),
        "color_count": Counter(),
        "coupe_count": Counter(),
        "forme_count": Counter(),
        "motif_count": Counter(),
        "collection_count": Counter(),
        "saison_count": Counter(),
        "genre_count": Counter(),
        "guide_count": Counter(),
        "category_count": Counter(),
        "material_by_type": defaultdict(Counter),
        "coupe_by_type": defaultdict(Counter),
        "price_by_material": defaultdict(list),
        "combos_type_mat": Counter(),
        "combos_type_col": Counter(),
        "combos_type_coupe": Counter(),
        "combos_type_coll": Counter(),
        "combos_mat_col": Counter(),
        "combos_type_genre": Counter(),
        "combos_type_genre_mat": Counter(),
        "combos_type_genre_col": Counter(),
        "combos_type_genre_coupe": Counter(),
        "taxonomy": defaultdict(set),
        "total": len(parsed_products),
        "total_in_stock": 0,
        "total_out_of_stock": 0,
    }
    for p in parsed_products:
        ptype = p["type"]
        is_in_stock = p.get("in_stock", True)  # fallback True pour WooCommerce
        results["type_count"][ptype] += 1
        results["taxonomy"]["Type de produit"].add(ptype)

        if is_in_stock:
            results["total_in_stock"] += 1
        else:
            results["total_out_of_stock"] += 1

        for mat in p["materials"]:
            results["materials_count"][mat] += 1
            results["material_by_type"][mat][ptype] += 1
            results["taxonomy"]["Matières"].add(mat)
            if p["price"]:
                results["price_by_material"][mat].append(p["price"])
            # Combos: uniquement les produits en stock
            if is_in_stock:
                results["combos_type_mat"][f"{ptype.lower()} {mat.lower()}"] += 1

        for col in p["colors"]:
            results["color_count"][col] += 1
            results["taxonomy"]["Couleurs"].add(col)
            if is_in_stock:
                results["combos_type_col"][f"{ptype.lower()} {col.lower()}"] += 1

        for coupe in p["coupes"]:
            results["coupe_count"][coupe] += 1
            results["coupe_by_type"][coupe][ptype] += 1
            results["taxonomy"]["Coupes"].add(coupe)
            if is_in_stock:
                results["combos_type_coupe"][f"{ptype.lower()} {coupe.lower()}"] += 1

        for forme in p["formes"]:
            results["forme_count"][forme] += 1
            results["taxonomy"]["Formes"].add(forme)

        for motif in p.get("motifs", []):
            results["motif_count"][motif] += 1
            results["taxonomy"]["Motifs"].add(motif)

        for coll in p["collections"]:
            results["collection_count"][coll] += 1
            results["taxonomy"]["Collections"].add(coll)
            if is_in_stock:
                results["combos_type_coll"][f"{ptype.lower()} {coll.lower()}"] += 1

        for saison in p.get("saisons", []):
            results["saison_count"][saison] += 1
            results["taxonomy"]["Saisons"].add(saison)

        for guide in p.get("guides", []):
            results["guide_count"][guide] += 1
            results["taxonomy"]["Guides de tailles"].add(guide)

        for cat in p.get("categories", []):
            results["category_count"][cat] += 1
            results["taxonomy"]["Catégories principales"].add(cat)

        # Triple combo: type + matière + couleur (uniquement en stock)
        if is_in_stock:
            for mat in p["materials"]:
                for col in p["colors"]:
                    results["combos_mat_col"][f"{ptype.lower()} {mat.lower()} {col.lower()}"] += 1

        # Detect genre: explicit tag OR inferred from collections/tags
        genres = list(p.get("genre", []))
        if not genres:
            # Infer from collections or raw tags
            all_text = " ".join(str(c).lower() for c in p.get("collections", []))
            all_text += " " + " ".join(str(t).lower() for t in p.get("tags", []))
            if "homme" in all_text and "femme" not in all_text:
                genres = ["Homme"]
            elif "femme" in all_text and "homme" not in all_text:
                genres = ["Femme"]
            elif "homme" in all_text and "femme" in all_text:
                genres = ["Homme", "Femme"]

        # Gender combos (uniquement en stock)
        if is_in_stock:
            for g in genres:
                g_lower = g.lower()
                results["genre_count"][g] += 1
                results["taxonomy"]["Genre"].add(g)
                results["combos_type_genre"][f"{ptype.lower()} {g_lower}"] += 1
                for mat in p["materials"]:
                    results["combos_type_genre_mat"][f"{ptype.lower()} {g_lower} {mat.lower()}"] += 1
                for col in p["colors"]:
                    results["combos_type_genre_col"][f"{ptype.lower()} {g_lower} {col.lower()}"] += 1
                for coupe in p["coupes"]:
                    results["combos_type_genre_coupe"][f"{ptype.lower()} {g_lower} {coupe.lower()}"] += 1

    return results


# ═════════════════════════════════════════════
# PARSEURS CSV AHREFS
# ═════════════════════════════════════════════

def parse_ahrefs_csv(uploaded_file):
    content = uploaded_file.read()
    try:
        text = content.decode("utf-16")
        sep = "\t"
    except (UnicodeDecodeError, UnicodeError):
        try:
            text = content.decode("utf-8-sig")
            sep = "\t" if "\t" in text[:500] else ","
        except UnicodeDecodeError:
            text = content.decode("latin-1")
            sep = "\t" if "\t" in text[:500] else ","
    df = pd.read_csv(StringIO(text), sep=sep)
    df.columns = [c.strip().strip('"') for c in df.columns]
    return df


def detect_ahrefs_type(df):
    cols = set(c.lower() for c in df.columns)
    if "top keyword" in cols or "top keyword: volume" in cols:
        return "top_pages"
    if "keyword" in cols and "volume" in cols:
        return "keywords"
    return "unknown"


# ═════════════════════════════════════════════
# MATCHING KEYWORDS VS PAGES
# ═════════════════════════════════════════════

def _url_words_match(kw_words, url_path):
    """Check if all keyword words appear in URL path (handling plurals)."""
    stop_words = {"en", "de", "du", "des", "le", "la", "les", "un", "une", "pour", "avec", "et", "homme", "femme"}
    words = [w for w in kw_words if w not in stop_words and len(w) > 1]
    if not words:
        return False
    for w in words:
        w_plural = w + "s" if not w.endswith("s") else w
        w_singular = w[:-1] if w.endswith("s") and len(w) > 2 else w
        if not (w in url_path or w_plural in url_path or w_singular in url_path):
            return False
    return True


def _is_dedicated_url(kw_words, url_path):
    """Check if URL is dedicated (only contains keyword words) vs partial (has extra content)."""
    stop_words = {"en", "de", "du", "des", "le", "la", "les", "un", "une", "pour", "avec", "et", "homme", "femme"}
    url_content_words = set(url_path.replace("-", " ").split()) - stop_words - {"homme", "femme"}
    kw_set = set(w for w in kw_words if w not in stop_words and len(w) > 1)
    kw_all_forms = kw_set | {w + "s" for w in kw_set if not w.endswith("s")} | {w[:-1] for w in kw_set if w.endswith("s") and len(w) > 2}
    return url_content_words <= kw_all_forms


def match_keywords_to_pages(df_keywords, df_pages, combos_with_materials, combos_counters=None, combos_category=None, df_internal=None):
    rows = []

    # Build page index by top keyword
    page_index = {}
    if df_pages is not None and not df_pages.empty:
        for _, row in df_pages.iterrows():
            kw = str(row.get("Top keyword", "")).strip().lower()
            url = str(row.get("URL", ""))
            if kw:
                page_index[kw] = {
                    "url": url,
                    "position": row.get("Top keyword: Position", ""),
                    "traffic": row.get("Traffic", 0),
                    "keywords_count": row.get("Keywords", 0),
                    "top_keyword": row.get("Top keyword", ""),
                }

    # Build volume index
    vol_index = {}
    if df_keywords is not None and not df_keywords.empty:
        for _, row in df_keywords.iterrows():
            kw = str(row.get("Keyword", "")).strip().lower()
            if kw:
                vol_index[kw] = {
                    "volume": row.get("Volume", 0),
                    "kd": row.get("Difficulty", ""),
                    "cpc": row.get("CPC", ""),
                    "traffic_potential": row.get("Traffic potential", ""),
                }

    for combo_kw, materials in combos_with_materials.items():
        combo_lower = combo_kw.lower().strip()
        vol_data = vol_index.get(combo_lower, {})
        # Skip combos not in Ahrefs CSV or without volume
        if not vol_data or not vol_data.get("volume") or pd.isna(vol_data.get("volume", 0)):
            continue
        volume = vol_data.get("volume", "N/A")
        kd = vol_data.get("kd", "N/A")
        cpc = vol_data.get("cpc", "N/A")
        tp = vol_data.get("traffic_potential", "N/A")
        kw_words = combo_lower.split()

        # ── Étape 1 : Page consacrée (matching par URL/slug) ──
        page_consacree = ""
        all_url_sources = []
        if df_pages is not None and not df_pages.empty:
            _pg_url_col = next((c for c in df_pages.columns if c.lower() in ("url", "address", "page url", "current url")), df_pages.columns[0])
            all_url_sources.extend(df_pages[_pg_url_col].dropna().tolist())
        if df_internal is not None and not df_internal.empty:
            _int_url_col = next((c for c in df_internal.columns if c.lower() in ("url", "address", "page url", "current url")), df_internal.columns[0])
            all_url_sources.extend(df_internal[_int_url_col].dropna().tolist())
        slug = combo_lower.replace(" ", "-")
        for src_url in all_url_sources:
            url = str(src_url).lower()
            url_path = url.split("/")[-1] if "/" in url else url
            if slug in url:
                page_consacree = str(src_url)
                break
            if _url_words_match(kw_words, url_path) and _is_dedicated_url(kw_words, url_path):
                page_consacree = str(src_url)
                break

        # ── Étape 2 : Page positionnée (matching par Top keyword Ahrefs) ──
        page_positionnee = ""
        position = ""
        traffic = 0
        top_kw_page = ""
        nb_kw_page = 0
        page_data = page_index.get(combo_lower, {})
        if page_data:
            page_positionnee = page_data.get("url", "")
            position = page_data.get("position", "")
            traffic = page_data.get("traffic", 0)
            top_kw_page = page_data.get("top_keyword", "")
            nb_kw_page = page_data.get("keywords_count", 0)

        # ── Correspondance et action ──
        if page_consacree and page_positionnee:
            correspondance = "Page dédiée"
        elif page_consacree:
            correspondance = "Page dédiée non positionnée"
        elif page_positionnee:
            correspondance = "Page positionnée"
        else:
            correspondance = "Pas de page"

        if correspondance == "Pas de page":
            action = "Créer page"
        elif correspondance == "Page positionnée":
            try:
                pos_int = int(float(str(position)))
            except (ValueError, TypeError):
                pos_int = 99
            action = "Suivre" if pos_int <= 3 else ("Optimiser" if pos_int <= 10 else "Améliorer")
        elif correspondance == "Page dédiée non positionnée":
            action = "Indexer / Optimiser"
        else:
            try:
                pos_int = int(float(str(position)))
            except (ValueError, TypeError):
                pos_int = 99
            action = "Suivre" if pos_int <= 3 else ("Optimiser" if pos_int <= 10 else "Améliorer")

        # Lookup nb produits
        nb_produits = 0
        if combos_counters and combos_category:
            cat = combos_category.get(combo_kw, "")
            counter_map = {
                "Type + Matière": "type_mat", "Type + Couleur": "type_col",
                "Type + Coupe": "type_coupe", "Type + Collection": "type_coll",
                "Type + Matière + Couleur": "mat_col", "Type + Genre": "type_genre",
                "Type + Genre + Matière": "type_genre_mat", "Type + Genre + Couleur": "type_genre_col",
                "Type + Genre + Coupe": "type_genre_coupe",
            }
            counter_key = counter_map.get(cat, "")
            if counter_key:
                nb_produits = combos_counters.get(counter_key, {}).get(combo_lower, 0)

        rows.append({
            "Mot-clé": combo_kw,
            "Nb produits": nb_produits if nb_produits else None,
            "Volume": volume,
            "KD": kd,
            "Potentiel trafic": tp,
            "CPC (€)": cpc,
            "Correspondance": correspondance,
            "Page consacrée": page_consacree if page_consacree else None,
            "Page positionnée": page_positionnee if page_positionnee else None,
            "Position": position if position else None,
            "Trafic page": traffic if traffic else None,
            "Top KW page": top_kw_page if top_kw_page else None,
            "Nb KW page": nb_kw_page if nb_kw_page else None,
            "Action recommandée": action,
            "_matiere": ", ".join(materials) if materials else "",
            "Type combinaison": combos_category.get(combo_kw, "Autre") if combos_category else "Autre",
        })

    df = pd.DataFrame(rows)
    if not df.empty:
        import numpy as np
        vol_num = pd.to_numeric(df["Volume"], errors="coerce").fillna(0)
        prod_num = pd.to_numeric(df["Nb produits"], errors="coerce").fillna(0)
        cpc_num = pd.to_numeric(df["CPC (€)"], errors="coerce").fillna(0)
        max_vol = vol_num.max() if vol_num.max() > 0 else 1
        max_cpc = cpc_num.max() if cpc_num.max() > 0 else 1
        cpc_ratio = cpc_num / max_cpc

        # Seuil de produits basé sur le volume + CPC
        def get_seuil(vol, cpc_r):
            if vol > 20000: base = 20
            elif vol > 10000: base = 15
            elif vol > 5000: base = 10
            elif vol > 1000: base = 5
            else: base = 3
            return base * (1 + cpc_r)

        seuils = pd.Series([get_seuil(v, c) for v, c in zip(vol_num, cpc_ratio)])
        couverture = (prod_num / seuils).clip(upper=1)

        # Partie 1 : Volume pondéré par la couverture (70 pts)
        score_vol = (vol_num / max_vol) * couverture * 70

        # Partie 2 : Profondeur catalogue (30 pts) — log plafonné à 28, 0 si < 3
        score_prod = prod_num.apply(lambda x: 0 if x < 3 else min(np.log(x) / np.log(28) * 30, 30))

        score = (score_vol + score_prod).round(0).astype(int)
        df["Score priorité"] = score
        df = df.sort_values("Score priorité", ascending=False)
    return df


# ═════════════════════════════════════════════
# EXPORT EXCEL 10 ONGLETS
# ═════════════════════════════════════════════

def make_bar(count, max_count):
    """Generate a visual bar like ████████"""
    if not max_count or not count:
        return ""
    ratio = min(count / max_count, 1.0)
    bars = int(ratio * 20)
    return "█" * bars


def build_excel(results, df_matched, store_name, df_keywords=None):
    buffer = BytesIO()
    wb = Workbook()
    total = results["total"]

    # Build keyword set from Ahrefs CSV for filtering — only keywords WITH volume
    _ahrefs_keywords = set()
    if df_keywords is not None and not df_keywords.empty:
        for _, row in df_keywords.iterrows():
            kw = str(row.get("Keyword", "")).strip().lower()
            vol = row.get("Volume", 0)
            if kw and pd.notna(vol) and vol > 0:
                _ahrefs_keywords.add(kw)

    hf = PatternFill(start_color="1B2A4A", end_color="1B2A4A", fill_type="solid")
    hfont = Font(bold=True, color="FFFFFF", size=11)

    def style_header(ws, max_col, row=1):
        for c in range(1, max_col + 1):
            cell = ws.cell(row=row, column=c)
            cell.fill = hf
            cell.font = hfont
            cell.alignment = Alignment(horizontal="center")

    # 1. Par Matière
    ws = wb.active
    ws.title = "📊 Par Matière"
    ws.append(["Matière", "Nb produits", "% du catalogue", "Prix moyen (€)", "Prix min (€)", "Prix max (€)"])
    style_header(ws, 6)
    for mat, count in results["materials_count"].most_common():
        prices = results["price_by_material"].get(mat, [])
        ws.append([mat, count, f"{count/total*100:.1f}%",
                    round(sum(prices)/len(prices), 2) if prices else "",
                    round(min(prices), 2) if prices else "",
                    round(max(prices), 2) if prices else ""])

    # 2. Par Type
    ws2 = wb.create_sheet("📦 Par Type")
    ws2.append(["Type de produit", "Nombre", "% du catalogue"])
    style_header(ws2, 3)
    for t, c in results["type_count"].most_common():
        ws2.append([t, c, f"{c/total*100:.1f}%"])

    # 3. Matière x Type
    ws3 = wb.create_sheet("🔀 Matière x Type")
    all_types = sorted(set(t for mat_c in results["material_by_type"].values() for t in mat_c))
    ws3.append(["Matière"] + all_types)
    style_header(ws3, 1 + len(all_types))
    for mat, tc in sorted(results["material_by_type"].items()):
        ws3.append([mat] + [tc.get(t, 0) for t in all_types])

    # 4. Par Couleur
    ws4 = wb.create_sheet("🎨 Par Couleur")
    ws4.append(["Couleur", "Nombre", "% du catalogue"])
    style_header(ws4, 3)
    for col, c in results["color_count"].most_common():
        ws4.append([col, c, f"{c/total*100:.1f}%"])

    # 5. Par Coupe
    ws5 = wb.create_sheet("📐 Par Coupe")
    ws5.append(["Coupe", "Nb produits", "% du catalogue"])
    style_header(ws5, 3)
    for cp, c in results["coupe_count"].most_common():
        ws5.append([cp, c, f"{c/total*100:.1f}%"])

    # 6. Tous les produits
    ws6 = wb.create_sheet("📋 Tous les produits")
    ws6.append(["Produit", "Type", "Matières", "Composition", "Couleurs", "Coupes", "Formes", "Prix (€)", "Ancien prix (€)", "En stock", "Tailles dispo", "Publié le", "URL"])
    style_header(ws6, 13)
    for p in results["products"]:
        stock_label = "✅" if p.get("in_stock", True) else "❌"
        tailles = f"{p.get('available_variants', '')}/{p.get('total_variants', '')}" if p.get("total_variants") else ""
        published = str(p.get("published_at", ""))[:10] if p.get("published_at") else ""
        ws6.append([p["title"], p["type"], p["materials_str"], p["composition"],
                     p["colors_str"], p["coupes_str"], p["formes_str"],
                     p["price"], p.get("compare_price", ""), stock_label, tailles, published, p["url"]])

    # 7. Taxonomie — 11 colonnes comme le ref
    ws7 = wb.create_sheet("🗂 Taxonomie")
    tax_cols_order = [
        "Type de produit", "Catégories principales", "Matières", "Couleurs",
        "Coupes", "Formes", "Motifs", "Collections", "Saisons", "Genre", "Guides de tailles"
    ]
    taxonomy = results["taxonomy"]
    # Ensure all expected columns exist
    for col in tax_cols_order:
        if col not in taxonomy:
            taxonomy[col] = set()
    max_len = max((len(v) for v in taxonomy.values()), default=0)
    ws7.append(tax_cols_order)
    style_header(ws7, len(tax_cols_order))
    for i in range(max_len):
        row = []
        for attr in tax_cols_order:
            vals = sorted(taxonomy.get(attr, set()))
            row.append(vals[i] if i < len(vals) else "")
        ws7.append(row)

    # 8. Mots-clés SEO — 8 colonnes, chaque colonne liste TOUTES ses combos indépendamment
    ws8 = wb.create_sheet("🔑 Mots-clés SEO")
    headers_seo = ["cat", "cat + sous cat", "cat + matières", "cat + couleurs",
                    "cat + coupes", "cat + formes", "cat + motifs", "cat + saisons",
                    "cat + genre", "cat + genre + matière", "cat + genre + couleur", "cat + genre + coupe"]
    ws8.append(headers_seo)
    style_header(ws8, len(headers_seo))

    types_list = sorted(results["type_count"].keys())
    cats_list = sorted(results["category_count"].keys()) if results.get("category_count") else []
    mats_list = sorted(results["materials_count"].keys())
    cols_list = sorted(results["color_count"].keys())
    coupes_list = sorted(results["coupe_count"].keys())
    formes_list = sorted(results["forme_count"].keys())
    motifs_list = sorted(results["motif_count"].keys()) if results.get("motif_count") else []
    saisons_list = sorted(results["saison_count"].keys()) if results.get("saison_count") else []

    # Build each column independently: only combos present in Ahrefs CSV
    def build_combos(attr_list):
        combos = []
        for t in types_list:
            for attr in attr_list:
                combo = f"{t.lower()} {attr.lower()}"
                if not _ahrefs_keywords or combo in _ahrefs_keywords:
                    combos.append(combo)
        return combos

    col_cat = types_list
    col_sous_cat = build_combos(cats_list)
    col_matieres = build_combos(mats_list)
    col_couleurs = build_combos(cols_list)
    col_coupes = build_combos(coupes_list)
    col_formes = build_combos(formes_list)
    col_motifs = build_combos(motifs_list)
    col_saisons = build_combos(saisons_list)

    genres_list = sorted(results["genre_count"].keys()) if results.get("genre_count") else []

    def build_genre_combos(attr_list=None):
        """Build type x genre (x attr) combos — only those in Ahrefs CSV."""
        combos = []
        for t in types_list:
            for g in genres_list:
                if attr_list:
                    for attr in attr_list:
                        combo = f"{t.lower()} {g.lower()} {attr.lower()}"
                        if not _ahrefs_keywords or combo in _ahrefs_keywords:
                            combos.append(combo)
                else:
                    combo = f"{t.lower()} {g.lower()}"
                    if not _ahrefs_keywords or combo in _ahrefs_keywords:
                        combos.append(combo)
        return combos

    col_genre = build_genre_combos()
    col_genre_mat = build_genre_combos(mats_list)
    col_genre_col = build_genre_combos(cols_list)
    col_genre_coupe = build_genre_combos(coupes_list)

    all_columns = [col_cat, col_sous_cat, col_matieres, col_couleurs,
                    col_coupes, col_formes, col_motifs, col_saisons,
                    col_genre, col_genre_mat, col_genre_col, col_genre_coupe]
    max_rows = max(len(c) for c in all_columns) if all_columns else 0

    for i in range(max_rows):
        row = []
        for col_data in all_columns:
            row.append(col_data[i] if i < len(col_data) else "")
        ws8.append(row)

    # 9. Top Combinaisons — with visual bar + Ahrefs data (5 sections)
    ws9 = wb.create_sheet("🏆 Top Combinaisons")

    # Build vol_index for Ahrefs enrichment
    vol_index = {}
    if df_keywords is not None and not df_keywords.empty:
        for _, row in df_keywords.iterrows():
            kw = str(row.get("Keyword", "")).strip().lower()
            if kw:
                vol_index[kw] = {
                    "volume": row.get("Volume", ""),
                    "kd": row.get("Difficulty", ""),
                    "tp": row.get("Traffic potential", ""),
                    "cpc": row.get("CPC", ""),
                }

    def write_combo_section(ws, title, example, counter, start_row):
        """Write a combo section: title, header, data rows, total, blank line."""
        ws.append([title, None, example])
        ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=3)
        ws.cell(start_row, 1).font = Font(bold=True, size=13)
        ws.append(["Combinaison", "Nb produits", "% du total", None, "Volume", "KD", "Potentiel trafic", "CPC (€)"])
        style_header(ws, 8, row=start_row + 1)
        combos_sorted = [(c, n) for c, n in counter.most_common() if not _ahrefs_keywords or c.lower() in _ahrefs_keywords]
        max_c = combos_sorted[0][1] if combos_sorted else 1
        for combo, count in combos_sorted:
            bar = make_bar(count, max_c)
            vd = vol_index.get(combo.lower(), {})
            ws.append([
                combo, count, f"{count/total*100:.1f}%", bar,
                vd.get("volume", ""), vd.get("kd", ""),
                vd.get("tp", ""), vd.get("cpc", ""),
            ])
        ws.append([f"Total combinaisons uniques : {len(combos_sorted)}"])
        ws.append([])  # blank line
        return start_row + len(combos_sorted) + 4  # title + header + data + total + blank

    current_row = 1
    if results["combos_type_mat"]:
        current_row = write_combo_section(ws9, "🧵 Type + Matière", "Ex: chemise coton, pull laine...",
                                           results["combos_type_mat"], current_row)
    if results["combos_type_col"]:
        current_row = write_combo_section(ws9, "🎨 Type + Couleur", "Ex: chemise bleu, pantalon beige...",
                                           results["combos_type_col"], current_row)
    if results["combos_type_coupe"]:
        current_row = write_combo_section(ws9, "📐 Type + Coupe", "Ex: chemise slim, pantalon ajustée...",
                                           results["combos_type_coupe"], current_row)
    if results.get("combos_type_coll"):
        current_row = write_combo_section(ws9, "📁 Type + Collection", "Ex: chemise casual, pull mailles...",
                                           results["combos_type_coll"], current_row)
    if results.get("combos_mat_col"):
        current_row = write_combo_section(ws9, "🔗 Type + Matière + Couleur", "Ex: chemise coton bleu...",
                                           results["combos_mat_col"], current_row)
    if results.get("combos_type_genre"):
        current_row = write_combo_section(ws9, "👤 Type + Genre", "Ex: chemise homme, pull femme...",
                                           results["combos_type_genre"], current_row)
    if results.get("combos_type_genre_mat"):
        current_row = write_combo_section(ws9, "👤🧵 Type + Genre + Matière", "Ex: chemise homme coton, pull homme laine...",
                                           results["combos_type_genre_mat"], current_row)
    if results.get("combos_type_genre_col"):
        current_row = write_combo_section(ws9, "👤🎨 Type + Genre + Couleur", "Ex: chemise homme bleu, pull homme vert...",
                                           results["combos_type_genre_col"], current_row)
    if results.get("combos_type_genre_coupe"):
        current_row = write_combo_section(ws9, "👤📐 Type + Genre + Coupe", "Ex: chemise homme slim, pantalon homme ajustée...",
                                           results["combos_type_genre_coupe"], current_row)

    # 10. Requêtes vs Pages — matching ref format
    if df_matched is not None and not df_matched.empty:
        ws10 = wb.create_sheet("🔍 Requêtes vs Pages")
        cols_export = ["Mot-clé", "Type combinaison", "Nb produits", "Volume", "Score priorité", "KD", "Potentiel trafic", "CPC (€)",
                        "Correspondance", "Page consacrée", "Page positionnée", "Position", "Trafic page", "Top KW page", "Nb KW page", "Action recommandée"]
        cols_data = cols_export
        ws10.append(cols_export)
        style_header(ws10, len(cols_export))

        for _, row in df_matched.iterrows():
            ws10.append([row.get(c, "") for c in cols_data])

        # Color correspondance column
        corresp_colors = {
            "Pas de page": "E74C3C",
            "Page positionnée": "F1C40F",
            "Page dédiée non positionnée": "3498DB",
            "Page dédiée": "27AE60",
        }
        corresp_col_idx = cols_export.index("Correspondance") + 1
        action_col_idx = cols_export.index("Action recommandée") + 1
        score_col_idx = cols_export.index("Score priorité") + 1
        for row_idx in range(2, ws10.max_row + 1):
            cell = ws10.cell(row=row_idx, column=corresp_col_idx)
            color = corresp_colors.get(str(cell.value), "FFFFFF")
            if color != "FFFFFF":
                cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                if color in ("E74C3C", "27AE60"):
                    cell.font = Font(color="FFFFFF", bold=True)

            # Color action column
            action_cell = ws10.cell(row=row_idx, column=action_col_idx)
            action_colors = {"Créer page": "E74C3C", "Optimiser": "F1C40F", "Améliorer": "E67E22", "Suivre": "27AE60", "Indexer / Optimiser": "3498DB"}
            ac = action_colors.get(str(action_cell.value), "FFFFFF")
            if ac != "FFFFFF":
                action_cell.fill = PatternFill(start_color=ac, end_color=ac, fill_type="solid")
                if ac in ("E74C3C", "27AE60"):
                    action_cell.font = Font(color="FFFFFF", bold=True)

            # Color score column (gradient: red < 30, orange 30-60, green > 60)
            score_cell = ws10.cell(row=row_idx, column=score_col_idx)
            try:
                sv = int(float(str(score_cell.value or 0)))
            except (ValueError, TypeError):
                sv = 0
            if sv >= 60:
                score_cell.fill = PatternFill(start_color="27AE60", end_color="27AE60", fill_type="solid")
                score_cell.font = Font(color="FFFFFF", bold=True)
            elif sv >= 30:
                score_cell.fill = PatternFill(start_color="F1C40F", end_color="F1C40F", fill_type="solid")
                score_cell.font = Font(bold=True)
            elif sv > 0:
                score_cell.fill = PatternFill(start_color="E67E22", end_color="E67E22", fill_type="solid")
                score_cell.font = Font(color="FFFFFF")

    for sheet in wb.worksheets:
        for c in range(1, sheet.max_column + 1):
            best = 0
            for r in range(1, min(sheet.max_row + 1, 500)):
                val = sheet.cell(row=r, column=c).value
                if val is not None:
                    s = str(val)
                    # Emoji/unicode chars take ~2 char widths in Excel
                    v = sum(2 if ord(ch) > 0xFFFF else 1 for ch in s)
                    if v > best:
                        best = v
            sheet.column_dimensions[get_column_letter(c)].width = min(max(best + 4, 12), 65)

    wb.save(buffer)
    return buffer.getvalue()


# ═════════════════════════════════════════════
# SIDEBAR — INPUTS
# ═════════════════════════════════════════════

with st.sidebar:
    st.header("📥 Sources de données")

    st.markdown("**1. Site e-commerce**")
    store_url = st.text_input("URL du site", placeholder="www.balibaris.com",
                               help="Shopify ou WooCommerce")
    cms_choice = st.selectbox("CMS", ["Auto-détecter", "Shopify", "WooCommerce"])

    if st.button("🚀 Scraper", type="primary", use_container_width=True):
        if store_url:
            progress = st.progress(0, "Scraping...")
            domain = store_url.replace("https://", "").replace("http://", "").strip("/")
            if cms_choice == "WooCommerce":
                raw = fetch_woo_products(domain, progress)
                cms = "woocommerce"
            elif cms_choice == "Shopify":
                raw = fetch_shopify_products(domain, progress)
                cms = "shopify"
            else:
                raw = fetch_shopify_products(domain, progress)
                cms = "shopify"
                if not raw:
                    st.info("Tentative WooCommerce...")
                    raw = fetch_woo_products(domain, progress)
                    cms = "woocommerce"
            if raw:
                store_clean = domain.replace("www.", "")
                parser = parse_shopify_product if cms == "shopify" else parse_woo_product
                parsed = [parser(p, domain) for p in raw]
                st.session_state["ac_results"] = analyze_parsed_products(parsed)
                st.session_state["ac_store"] = store_clean
                st.success(f"✅ {len(raw)} produits")
            else:
                st.error("❌ Aucun produit trouvé")

    if "ac_results" in st.session_state:
        r = st.session_state["ac_results"]
        in_stock = r.get("total_in_stock", r["total"])
        out_stock = r.get("total_out_of_stock", 0)
        st.success(f"✅ {r['total']} produits chargés — 📦 {in_stock} en stock, ❌ {out_stock} en rupture")

    st.divider()

    st.markdown("**2. Mots-clés Ahrefs**")
    uploaded_kw = st.file_uploader("CSV Keywords", type=["csv"], key="ac_kw",
                                    help="Export Ahrefs Keywords Explorer")

    st.markdown("**3. Top Pages Ahrefs**")
    uploaded_pages = st.file_uploader("CSV Top Pages", type=["csv"], key="ac_pages",
                                       help="Export Ahrefs Top Pages")

    st.markdown("**4. Pages internes (optionnel)**")
    uploaded_internal = st.file_uploader("CSV/HTML pages internes", type=["csv", "html"], key="ac_internal",
                                          help="Export Screaming Frog ou sitemap — liste de toutes les URLs du site")

    df_keywords = None
    df_pages = None
    df_internal = None

    if uploaded_kw:
        df_keywords = parse_ahrefs_csv(uploaded_kw)
        st.session_state["ac_df_keywords"] = df_keywords
        st.success(f"✅ {len(df_keywords)} mots-clés")
    elif "ac_df_keywords" in st.session_state:
        df_keywords = st.session_state["ac_df_keywords"]

    if uploaded_pages:
        df_pages = parse_ahrefs_csv(uploaded_pages)
        st.session_state["ac_df_pages"] = df_pages
        st.success(f"✅ {len(df_pages)} pages")
    elif "ac_df_pages" in st.session_state:
        df_pages = st.session_state["ac_df_pages"]

    if uploaded_internal:
        import re as _re
        name = uploaded_internal.name.lower()
        if name.endswith(".csv"):
            df_internal = parse_ahrefs_csv(uploaded_internal)
        elif name.endswith(".html"):
            content = uploaded_internal.read().decode("utf-8", errors="ignore")
            urls = _re.findall(r'https?://[^\s<>"\']+', content)
            df_internal = pd.DataFrame({"URL": list(set(urls))})
        if df_internal is not None and not df_internal.empty:
            st.session_state["ac_df_internal"] = df_internal
            st.success(f"✅ {len(df_internal)} pages internes")
    elif "ac_df_internal" in st.session_state:
        df_internal = st.session_state["ac_df_internal"]


# ═════════════════════════════════════════════
# ZONE PRINCIPALE
# ═════════════════════════════════════════════

has_products = "ac_results" in st.session_state
has_ahrefs = df_keywords is not None or df_pages is not None

if not has_products and not has_ahrefs:
    st.info("👈 Utilise la sidebar pour charger tes données : URL du site + exports Ahrefs.")
    st.stop()

df_matched = pd.DataFrame()
combos_with_materials = {}
combos_category = {}  # tracks which category each combo belongs to

if has_products:
    results = st.session_state["ac_results"]
    store = st.session_state.get("ac_store", "site")
    total = results["total"]

    for combo_key in results.get("combos_type_mat", {}):
        parts = combo_key.split()
        if len(parts) >= 2:
            mat = " ".join(parts[1:]).capitalize()
            combos_with_materials[combo_key] = [mat]
        combos_category[combo_key] = "Type + Matière"
    for combo_key in results.get("combos_type_col", {}):
        if combo_key not in combos_with_materials:
            combos_with_materials[combo_key] = []
        combos_category[combo_key] = "Type + Couleur"
    for combo_key in results.get("combos_type_coupe", {}):
        if combo_key not in combos_with_materials:
            combos_with_materials[combo_key] = []
        combos_category[combo_key] = "Type + Coupe"
    for combo_key in results.get("combos_type_coll", {}):
        if combo_key not in combos_with_materials:
            combos_with_materials[combo_key] = []
        combos_category[combo_key] = "Type + Collection"
    for combo_key in results.get("combos_mat_col", {}):
        if combo_key not in combos_with_materials:
            parts = combo_key.split()
            mat = parts[1].capitalize() if len(parts) >= 2 else ""
            combos_with_materials[combo_key] = [mat] if mat else []
        combos_category[combo_key] = "Type + Matière + Couleur"
    for combo_key in results.get("combos_type_genre", {}):
        if combo_key not in combos_with_materials:
            combos_with_materials[combo_key] = []
        combos_category[combo_key] = "Type + Genre"
    for combo_key in results.get("combos_type_genre_mat", {}):
        if combo_key not in combos_with_materials:
            parts = combo_key.split()
            mat = parts[2].capitalize() if len(parts) >= 3 else ""
            combos_with_materials[combo_key] = [mat] if mat else []
        combos_category[combo_key] = "Type + Genre + Matière"
    for combo_key in results.get("combos_type_genre_col", {}):
        if combo_key not in combos_with_materials:
            combos_with_materials[combo_key] = []
        combos_category[combo_key] = "Type + Genre + Couleur"
    for combo_key in results.get("combos_type_genre_coupe", {}):
        if combo_key not in combos_with_materials:
            combos_with_materials[combo_key] = []
        combos_category[combo_key] = "Type + Genre + Coupe"

if not combos_with_materials and df_keywords is not None:
    for _, row in df_keywords.iterrows():
        kw = str(row.get("Keyword", ""))
        combos_with_materials[kw] = []

if has_ahrefs and combos_with_materials:
    combos_counters = {}
    if has_products:
        combos_counters = {
            "type_mat": results.get("combos_type_mat", {}),
            "type_col": results.get("combos_type_col", {}),
            "type_coupe": results.get("combos_type_coupe", {}),
            "type_coll": results.get("combos_type_coll", {}),
            "mat_col": results.get("combos_mat_col", {}),
            "type_genre": results.get("combos_type_genre", {}),
            "type_genre_mat": results.get("combos_type_genre_mat", {}),
            "type_genre_col": results.get("combos_type_genre_col", {}),
            "type_genre_coupe": results.get("combos_type_genre_coupe", {}),
        }
    df_matched = match_keywords_to_pages(df_keywords, df_pages, combos_with_materials, combos_counters, combos_category, df_internal=df_internal)


# ── TABS ──
tabs_list = []
if has_products:
    tabs_list += ["📦 Catalogue", "🔀 Croisements", "🔑 Mots-clés SEO"]
if has_ahrefs:
    tabs_list += ["📊 Positions & Gaps", "🎯 Opportunités"]
tabs_list.append("📥 Export Excel")

tabs = st.tabs(tabs_list)
tab_idx = 0

# ── TAB CATALOGUE ──
if has_products:
    with tabs[tab_idx]:
        c1, c2, c3, c4, c5, c6 = st.columns(6)
        c1.metric("Produits", total)
        c2.metric("Types", len(results["type_count"]))
        c3.metric("Matières", len(results["materials_count"]))
        c4.metric("Couleurs", len(results["color_count"]))
        c5.metric("Coupes", len(results["coupe_count"]))
        c6.metric("👤 Genres", len(results.get("genre_count", {})))

        import plotly.express as px

        sub_t1, sub_t2, sub_t3, sub_t4, sub_t5 = st.tabs(["🧵 Matières", "📦 Types", "🎨 Couleurs", "📐 Coupes", "👤 Genres"])

        with sub_t1:
            if results["materials_count"]:
                df_mat = pd.DataFrame(results["materials_count"].most_common(), columns=["Matière", "Nb produits"])
                df_mat["% catalogue"] = (df_mat["Nb produits"] / total * 100).round(1)
                prix_data = []
                for mat in df_mat["Matière"]:
                    prices = results["price_by_material"].get(mat, [])
                    prix_data.append({
                        "Prix moyen": round(sum(prices)/len(prices), 2) if prices else None,
                        "Prix min": round(min(prices), 2) if prices else None,
                        "Prix max": round(max(prices), 2) if prices else None,
                    })
                df_mat = pd.concat([df_mat, pd.DataFrame(prix_data)], axis=1)
                col1, col2 = st.columns([2, 1])
                with col1:
                    st.dataframe(df_mat, use_container_width=True, hide_index=True)
                with col2:
                    fig = px.pie(df_mat.head(15), values="Nb produits", names="Matière", title="Top 15 matières")
                    fig.update_layout(height=400)
                    st.plotly_chart(fig, use_container_width=True)

        with sub_t2:
            if results["type_count"]:
                df_type = pd.DataFrame(results["type_count"].most_common(), columns=["Type", "Nb produits"])
                df_type["% catalogue"] = (df_type["Nb produits"] / total * 100).round(1)
                col1, col2 = st.columns([2, 1])
                with col1:
                    st.dataframe(df_type, use_container_width=True, hide_index=True)
                with col2:
                    fig = px.bar(df_type.head(15), x="Type", y="Nb produits", title="Par type")
                    fig.update_layout(height=400)
                    st.plotly_chart(fig, use_container_width=True)

        with sub_t3:
            if results["color_count"]:
                df_col = pd.DataFrame(results["color_count"].most_common(), columns=["Couleur", "Nb produits"])
                df_col["% catalogue"] = (df_col["Nb produits"] / total * 100).round(1)
                col1, col2 = st.columns([2, 1])
                with col1:
                    st.dataframe(df_col, use_container_width=True, hide_index=True)
                with col2:
                    fig = px.bar(df_col.head(20), x="Couleur", y="Nb produits", title="Top 20 couleurs")
                    fig.update_layout(height=400)
                    st.plotly_chart(fig, use_container_width=True)

        with sub_t4:
            if results["coupe_count"]:
                df_coupe = pd.DataFrame(results["coupe_count"].most_common(), columns=["Coupe", "Nb produits"])
                df_coupe["% catalogue"] = (df_coupe["Nb produits"] / total * 100).round(1)
                col1, col2 = st.columns([2, 1])
                with col1:
                    st.dataframe(df_coupe, use_container_width=True, hide_index=True)
                with col2:
                    fig = px.pie(df_coupe, values="Nb produits", names="Coupe", title="Coupes")
                    fig.update_layout(height=400)
                    st.plotly_chart(fig, use_container_width=True)
            else:
                st.info("Aucune donnée de coupe trouvée")

        with sub_t5:
            if results.get("genre_count"):
                df_genre = pd.DataFrame(results["genre_count"].most_common(), columns=["Genre", "Nb produits"])
                df_genre["% catalogue"] = (df_genre["Nb produits"] / total * 100).round(1)
                col1, col2 = st.columns([2, 1])
                with col1:
                    st.dataframe(df_genre, use_container_width=True, hide_index=True)
                with col2:
                    fig = px.pie(df_genre, values="Nb produits", names="Genre", title="Répartition par genre")
                    fig.update_layout(height=400)
                    st.plotly_chart(fig, use_container_width=True)

                # Détail : combos Type + Genre
                if results.get("combos_type_genre"):
                    st.markdown("**Produits par Type × Genre**")
                    genre_combo_data = []
                    for combo, count in results["combos_type_genre"].most_common():
                        parts = combo.split()
                        if len(parts) >= 2:
                            genre_combo_data.append({"Type": parts[0].capitalize(), "Genre": parts[1].capitalize(), "Nb produits": count})
                    if genre_combo_data:
                        df_tg = pd.DataFrame(genre_combo_data)
                        pivot = df_tg.pivot_table(index="Type", columns="Genre", values="Nb produits", fill_value=0, aggfunc="sum")
                        st.dataframe(pivot, use_container_width=True)
            else:
                st.info("Aucune donnée de genre trouvée")

    tab_idx += 1

    # ── TAB CROISEMENTS ──
    with tabs[tab_idx]:
        import plotly.express as px

        st.markdown("#### Matière x Type de produit")
        if results["material_by_type"]:
            all_types = sorted(set(t for mat_c in results["material_by_type"].values() for t in mat_c))
            cross_data = []
            for mat, tc in sorted(results["material_by_type"].items()):
                row = {"Matière": mat}
                for t in all_types:
                    row[t] = tc.get(t, 0)
                cross_data.append(row)
            df_cross = pd.DataFrame(cross_data)
            st.dataframe(df_cross, use_container_width=True, hide_index=True)

            df_heat = df_cross.set_index("Matière")
            fig = px.imshow(df_heat.head(20), text_auto=True, aspect="auto",
                           title="Heatmap Matière x Type (top 20)")
            fig.update_layout(height=600)
            st.plotly_chart(fig, use_container_width=True)

        st.divider()
        st.markdown("#### Taxonomie complète")
        taxonomy = results["taxonomy"]
        if taxonomy:
            tax_cols_order = [
                "Type de produit", "Catégories principales", "Matières", "Couleurs",
                "Coupes", "Formes", "Motifs", "Collections", "Saisons", "Genre", "Guides de tailles"
            ]
            max_len = max((len(taxonomy.get(c, set())) for c in tax_cols_order), default=0)
            tax_data = {}
            for col_name in tax_cols_order:
                vals = sorted(taxonomy.get(col_name, set()))
                tax_data[col_name] = vals + [""] * (max_len - len(vals))
            st.dataframe(pd.DataFrame(tax_data), use_container_width=True, hide_index=True)

        st.divider()
        st.markdown("#### Liste complète des produits")
        df_all = pd.DataFrame([{
            "Produit": p["title"], "Type": p["type"], "Matières": p["materials_str"],
            "Couleurs": p["colors_str"], "Coupes": p["coupes_str"],
            "Prix": p["price"], "URL": p["url"],
        } for p in results["products"]])
        st.dataframe(df_all, use_container_width=True, hide_index=True,
                      column_config={"URL": st.column_config.LinkColumn("URL")})

    tab_idx += 1

    # ── TAB MOTS-CLÉS SEO ──
    with tabs[tab_idx]:
        st.markdown("#### Combinaisons SEO — toutes les combinaisons type + attribut")

        # ── COPIER / COLLER POUR AHREFS (en haut, bien visible) ──
        with st.expander("📋 **Exporter les combinaisons pour Ahrefs**", expanded=True):
            st.caption("Sélectionne les types de combinaisons, copie la liste et colle-la dans Ahrefs Keywords Explorer.")

            combo_sources = {
                "🧵 Type + Matière": results.get("combos_type_mat", {}),
                "🎨 Type + Couleur": results.get("combos_type_col", {}),
                "📐 Type + Coupe": results.get("combos_type_coupe", {}),
                "📁 Type + Collection": results.get("combos_type_coll", {}),
                "🔗 Type + Matière + Couleur": results.get("combos_mat_col", {}),
                "👤 Type + Genre": results.get("combos_type_genre", {}),
                "👤🧵 Type + Genre + Matière": results.get("combos_type_genre_mat", {}),
                "👤🎨 Type + Genre + Couleur": results.get("combos_type_genre_col", {}),
                "👤📐 Type + Genre + Coupe": results.get("combos_type_genre_coupe", {}),
            }
            combo_sources = {k: v for k, v in combo_sources.items() if v}

            sel_sources = st.multiselect(
                "Types de combinaisons à inclure",
                options=list(combo_sources.keys()),
                default=list(combo_sources.keys()),
                key="ahrefs_combo_select"
            )

            all_kws = []
            for src in sel_sources:
                for kw, nb in combo_sources.get(src, {}).items():
                    if nb > 0 and kw not in all_kws:
                        all_kws.append(kw)
            all_kws.sort()

            col_info, col_dl = st.columns([3, 1])
            with col_info:
                st.metric("Mots-clés à exporter", len(all_kws))
            with col_dl:
                if all_kws:
                    csv_content = "\n".join(all_kws)
                    st.download_button(
                        "⬇️ Télécharger en CSV",
                        data=csv_content,
                        file_name=f"combinaisons_ahrefs_{len(all_kws)}.csv",
                        mime="text/csv",
                        key="dl_ahrefs_csv"
                    )

            if all_kws:
                kw_text = "\n".join(all_kws)
                st.text_area(
                    "Liste des mots-clés (un par ligne — Ctrl+A puis Ctrl+C pour copier)",
                    value=kw_text,
                    height=300,
                    key="ahrefs_kw_list"
                )

        st.divider()

        if results["combos_type_mat"]:
            st.markdown("**🏆 Top Combinaisons Type + Matière**")
            combos_sorted = results["combos_type_mat"].most_common(50)
            max_c = combos_sorted[0][1] if combos_sorted else 1
            combo_rows = []
            for combo, count in combos_sorted:
                bar = make_bar(count, max_c)
                combo_rows.append({"Combinaison": combo, "Nb produits": count,
                                    "% du total": f"{count/total*100:.1f}%", "": bar})
            st.dataframe(pd.DataFrame(combo_rows), use_container_width=True, hide_index=True)

        if results["combos_type_col"]:
            st.markdown("**🎨 Type + Couleur** (top 30)")
            df_cc = pd.DataFrame(results["combos_type_col"].most_common(30),
                                  columns=["Combinaison", "Nb produits"])
            st.dataframe(df_cc, use_container_width=True, hide_index=True)

        if results["combos_type_coupe"]:
            st.markdown("**📐 Type + Coupe** (top 30)")
            df_ccp = pd.DataFrame(results["combos_type_coupe"].most_common(30),
                                   columns=["Combinaison", "Nb produits"])
            st.dataframe(df_ccp, use_container_width=True, hide_index=True)

        if results.get("combos_type_genre"):
            st.markdown("**👤 Type + Genre** (top 30)")
            df_tg = pd.DataFrame(results["combos_type_genre"].most_common(30),
                                  columns=["Combinaison", "Nb produits"])
            st.dataframe(df_tg, use_container_width=True, hide_index=True)

        if results.get("combos_type_genre_mat"):
            st.markdown("**👤🧵 Type + Genre + Matière** (top 30)")
            df_tgm = pd.DataFrame(results["combos_type_genre_mat"].most_common(30),
                                   columns=["Combinaison", "Nb produits"])
            st.dataframe(df_tgm, use_container_width=True, hide_index=True)

        if results.get("combos_type_genre_col"):
            st.markdown("**👤🎨 Type + Genre + Couleur** (top 30)")
            df_tgc = pd.DataFrame(results["combos_type_genre_col"].most_common(30),
                                   columns=["Combinaison", "Nb produits"])
            st.dataframe(df_tgc, use_container_width=True, hide_index=True)

        if results.get("combos_type_genre_coupe"):
            st.markdown("**👤📐 Type + Genre + Coupe** (top 30)")
            df_tgcp = pd.DataFrame(results["combos_type_genre_coupe"].most_common(30),
                                    columns=["Combinaison", "Nb produits"])
            st.dataframe(df_tgcp, use_container_width=True, hide_index=True)

    tab_idx += 1

# ── TAB POSITIONS & GAPS ──
if has_ahrefs:
    with tabs[tab_idx]:
        if df_matched.empty:
            st.info("Charge au moins un CSV Ahrefs + scrape un site pour voir les positions.")
        else:
            import plotly.express as px

            st.markdown("### 🎛️ Filtres")
            # Extract type from keyword (first word)
            df_matched["_type"] = df_matched["Mot-clé"].apply(lambda x: str(x).split()[0].capitalize() if x else "")
            all_types = sorted(df_matched["_type"].unique())
            all_materials = sorted(set(m for mats in combos_with_materials.values() for m in mats if m))
            all_combo_cats = sorted(df_matched["Type combinaison"].unique())

            # Detect genre from keyword for filtering
            genre_keywords = {"homme", "femme"}
            df_matched["_genre"] = df_matched["Mot-clé"].apply(
                lambda x: next((w.capitalize() for w in str(x).lower().split() if w in genre_keywords), "Non genré"))
            all_genres_filter = sorted(df_matched["_genre"].unique())

            fc0, fc1, fc2, fc3, fc4, fc5 = st.columns(6)
            with fc0:
                sel_combo_cat = st.multiselect("🏷️ Combinaison", options=all_combo_cats, default=[])
            with fc1:
                sel_type = st.multiselect("📦 Type", options=all_types, default=[])
            with fc2:
                sel_mat = st.multiselect("🧵 Matière", options=all_materials, default=[])
            with fc3:
                sel_genre = st.multiselect("👤 Genre", options=all_genres_filter, default=[])
            with fc4:
                all_corresp = sorted(df_matched["Correspondance"].unique())
                sel_corresp = st.multiselect("📌 Correspondance", options=all_corresp, default=[])
            with fc5:
                vol_min = st.number_input("🔢 Volume min", min_value=0, value=0, step=100)

            df_display = df_matched.copy()
            if sel_combo_cat:
                df_display = df_display[df_display["Type combinaison"].isin(sel_combo_cat)]
            if sel_type:
                df_display = df_display[df_display["_type"].isin(sel_type)]
            if sel_mat:
                df_display = df_display[df_display["_matiere"].apply(
                    lambda x: any(m.lower() in str(x).lower() for m in sel_mat))]
            if sel_genre:
                df_display = df_display[df_display["_genre"].isin(sel_genre)]
            if sel_corresp:
                df_display = df_display[df_display["Correspondance"].isin(sel_corresp)]
            if vol_min > 0:
                df_display = df_display[pd.to_numeric(df_display["Volume"], errors="coerce").fillna(0) >= vol_min]

            # KPIs
            total_kw = len(df_display)
            dedicated = len(df_display[df_display["Correspondance"] == "Page dédiée"])
            partial = len(df_display[df_display["Correspondance"] == "Page partielle"])
            no_page = len(df_display[df_display["Correspondance"] == "Pas de page"])
            with_page = dedicated + partial

            k1, k2, k3, k4, k5 = st.columns(5)
            k1.metric("Total", total_kw)
            k2.metric("🟢 Page dédiée", dedicated)
            k3.metric("🟡 Page partielle", partial)
            k4.metric("🔴 Pas de page", no_page)
            k5.metric("Couverture", f"{round(with_page/max(total_kw,1)*100)}%")

            # Charts
            ch1, ch2 = st.columns(2)
            with ch1:
                corresp_counts = df_display["Correspondance"].value_counts()
                fig = px.pie(values=corresp_counts.values, names=corresp_counts.index,
                            title="Correspondance",
                            color_discrete_map={
                                "Page dédiée": "#2ecc71", "Page partielle": "#f1c40f",
                                "Pas de page": "#e74c3c"})
                st.plotly_chart(fig, use_container_width=True)
            with ch2:
                action_counts = df_display["Action recommandée"].value_counts()
                fig2 = px.bar(x=action_counts.index, y=action_counts.values,
                              title="Actions recommandées",
                              labels={"x": "Action", "y": "Nb mots-clés"},
                              color=action_counts.index,
                              color_discrete_map={
                                  "Créer page": "#e74c3c", "Optimiser": "#f1c40f",
                                  "Améliorer": "#e67e22", "Suivre": "#2ecc71"})
                st.plotly_chart(fig2, use_container_width=True)

            # Table — hide internal _matiere column
            st.markdown("### 📋 Détail")
            with st.expander("ℹ️ **Comment est calculé le Score de priorité ?**", expanded=False):
                st.markdown("""
Le **Score de priorité** (0–100) identifie les mots-clés où investir en priorité. Il combine 2 facteurs :

| Facteur | Poids | Logique |
|---|---|---|
| 📈 **Volume pondéré** | 70% | Le volume est pondéré par votre **capacité à rivaliser** : plus le volume est élevé, plus il faut de produits. Le CPC augmente cette exigence (SERP concurrentielle). |
| 📦 **Profondeur catalogue** | 30% | 0 pts si < 3 produits. Puis courbe logarithmique plafonnée à 28 produits. |

**Formule :**
```
seuil = seuil_base(volume) × (1 + CPC/CPC_max)

seuil_base : >20K vol → 20 prod | >10K → 15 | >5K → 10 | >1K → 5 | ≤1K → 3

couverture = min(Nb_produits / seuil, 1)

Score = (Volume / Vol_max) × couverture × 70
      + log(Nb_produits) / log(28) × 30        [0 si < 3 produits]
```

**Exemples :**
| Mot-clé | Volume | CPC | Produits | Seuil | Couverture | Score |
|---|---|---|---|---|---|---|
| chemise homme | 29 000 | 0.40€ | 28 | 24 | 100% | **69** ✅ |
| manteau homme | 27 000 | 0.80€ | 3 | 36 | 8% | **13** ⚠️ |
| chino beige | 800 | 0.10€ | 13 | 3 | 100% | **22** |
| blouson cuir | 5 000 | 1.00€ | 5 | 20 | 25% | **15** ⚠️ |

**Lecture :** Un score élevé = fort volume + catalogue suffisant pour rivaliser → **priorité maximale**.
Un score bas = pas assez de produits pour la concurrence, ou peu de volume.
""")
            display_cols = [c for c in df_display.columns if not c.startswith("_")]
            st.dataframe(df_display[display_cols], use_container_width=True, hide_index=True,
                          column_config={
                              "Page consacrée": st.column_config.LinkColumn("Page consacrée"),
                              "Page positionnée": st.column_config.LinkColumn("Page positionnée"),
                              "Volume": st.column_config.NumberColumn("Volume", format="%d"),
                              "Trafic page": st.column_config.NumberColumn("Trafic page", format="%d"),
                          })

            # Vue par matière
            if all_materials:
                st.markdown("### 🧵 Vue détaillée par matière")
                mat_choice = st.selectbox("Matière", options=["Toutes"] + all_materials)
                if mat_choice != "Toutes":
                    df_mv = df_matched[df_matched["_matiere"].str.contains(mat_choice, case=False, na=False)]
                    covered = len(df_mv[df_mv["Correspondance"] != "Pas de page"])
                    not_covered = len(df_mv[df_mv["Correspondance"] == "Pas de page"])
                    m1, m2, m3 = st.columns(3)
                    m1.metric(f"Mots-clés '{mat_choice}'", len(df_mv))
                    m2.metric("✅ Couverts", covered)
                    m3.metric("🔴 Non couverts", not_covered)
                    mv_cols = [c for c in df_mv.columns if not c.startswith("_")]
                    st.dataframe(df_mv[mv_cols], use_container_width=True, hide_index=True)

    tab_idx += 1

    # ── TAB OPPORTUNITÉS ──
    with tabs[tab_idx]:
        if df_matched.empty:
            st.info("Pas encore de données croisées.")
        else:
            st.markdown("### 🎯 Top Opportunités — mots-clés sans page dédiée")
            df_opps = df_matched[df_matched["Correspondance"] == "Pas de page"].copy()
            df_opps["_vol"] = pd.to_numeric(df_opps["Volume"], errors="coerce").fillna(0)
            df_opps = df_opps.sort_values("_vol", ascending=False).drop(columns="_vol")

            if not df_opps.empty:
                st.metric("Nombre d'opportunités", len(df_opps))
                opps_cols = [c for c in df_opps.columns if not c.startswith("_")]
                st.dataframe(df_opps[opps_cols], use_container_width=True, hide_index=True,
                              column_config={"Volume": st.column_config.NumberColumn("Volume", format="%d")})
            else:
                st.success("🎉 Toutes les combinaisons sont couvertes !")

    tab_idx += 1

# ── TAB EXPORT ──
with tabs[tab_idx]:
    st.markdown("### 📥 Export Excel complet — 10 onglets")

    if has_products:
        st.markdown("""
        **Onglets inclus :**
        1. 📊 Par Matière  2. 📦 Par Type  3. 🔀 Matière x Type  4. 🎨 Par Couleur
        5. 📐 Par Coupe  6. 📋 Tous les produits  7. 🗂 Taxonomie (11 colonnes)
        8. 🔑 Mots-clés SEO (8 colonnes)  9. 🏆 Top Combinaisons (+ Volume/KD/CPC)
        10. 🔍 Requêtes vs Pages (Correspondance + Action)
        """)

        if st.button("💾 Générer le fichier Excel", type="primary", use_container_width=True):
            excel_data = build_excel(results, df_matched if not df_matched.empty else None, store, df_keywords=df_keywords)
            ts = datetime.now().strftime("%Y%m%d_%H%M")
            st.download_button(
                "⬇️ Télécharger l'Excel",
                data=excel_data,
                file_name=f"analyse_{store}_{ts}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )
    elif has_ahrefs and not df_matched.empty:
        if st.button("💾 Exporter les positions", type="primary", use_container_width=True):
            buffer = BytesIO()
            with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
                exp_cols = [c for c in df_matched.columns if not c.startswith("_")]
                df_matched[exp_cols].to_excel(writer, sheet_name="Positions", index=False)
            ts = datetime.now().strftime("%Y%m%d_%H%M")
            st.download_button(
                "⬇️ Télécharger",
                data=buffer.getvalue(),
                file_name=f"positions_{ts}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )
    else:
        st.info("Charge des données pour pouvoir exporter.")
