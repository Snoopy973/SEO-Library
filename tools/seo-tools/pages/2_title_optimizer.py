import streamlit as st
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from io import BytesIO
import time
from datetime import datetime


st.title("📊 Outil de Ré-optimisation de Contenus SEO")
st.markdown("Générez un fichier d'analyse pour identifier les pages en perte de position, clics et CTR.")

# Session state pour les noms de colonnes
if 'column_names' not in st.session_state:
    st.session_state.column_names = {
        # GSC extraction
        'gsc_page': 'page',
        'gsc_clicks': 'clicks',
        'gsc_impressions': 'impressions',
        'gsc_ctr': 'ctr',
        'gsc_position': 'position',
        'gsc_start_date': 'start_date',
        'gsc_end_date': 'end_date',
        # GSC consolidation
        'consolidation_page': 'Page',
        'consolidation_keywords': 'Mots clés',
        # Ahrefs top pages
        'ahrefs_url': 'URL',
        'ahrefs_current_top_keyword': 'Current top keyword',
        'ahrefs_previous_position': 'Previous top keyword: Position',
        'ahrefs_current_position': 'Current top keyword: Position',
    }

# Section dépliante pour les noms de colonnes
with st.expander("⚙️ Configuration des noms de colonnes", expanded=False):
    st.markdown("**Modifiez les noms des colonnes si nécessaire (GSC et Ahrefs changent parfois les noms)**")
    
    col1, col2 = st.columns(2)
    
    with col1:
        st.markdown("##### Fichiers GSC (extraction par page)")
        st.session_state.column_names['gsc_page'] = st.text_input("Colonne Page", value=st.session_state.column_names['gsc_page'])
        st.session_state.column_names['gsc_clicks'] = st.text_input("Colonne Clics", value=st.session_state.column_names['gsc_clicks'])
        st.session_state.column_names['gsc_impressions'] = st.text_input("Colonne Impressions", value=st.session_state.column_names['gsc_impressions'])
        st.session_state.column_names['gsc_ctr'] = st.text_input("Colonne CTR", value=st.session_state.column_names['gsc_ctr'])
        st.session_state.column_names['gsc_position'] = st.text_input("Colonne Position", value=st.session_state.column_names['gsc_position'])
        st.session_state.column_names['gsc_start_date'] = st.text_input("Colonne Date début", value=st.session_state.column_names['gsc_start_date'])
        st.session_state.column_names['gsc_end_date'] = st.text_input("Colonne Date fin", value=st.session_state.column_names['gsc_end_date'])
    
    with col2:
        st.markdown("##### Fichier Consolidation GSC")
        st.session_state.column_names['consolidation_page'] = st.text_input("Colonne Page (consolidation)", value=st.session_state.column_names['consolidation_page'])
        st.session_state.column_names['consolidation_keywords'] = st.text_input("Colonne Mots-clés", value=st.session_state.column_names['consolidation_keywords'])
        
        st.markdown("##### Fichier Top Pages Ahrefs")
        st.session_state.column_names['ahrefs_url'] = st.text_input("Colonne URL", value=st.session_state.column_names['ahrefs_url'])
        st.session_state.column_names['ahrefs_current_top_keyword'] = st.text_input("Colonne Top Keyword actuel", value=st.session_state.column_names['ahrefs_current_top_keyword'])
        st.session_state.column_names['ahrefs_previous_position'] = st.text_input("Colonne Position précédente", value=st.session_state.column_names['ahrefs_previous_position'])
        st.session_state.column_names['ahrefs_current_position'] = st.text_input("Colonne Position actuelle", value=st.session_state.column_names['ahrefs_current_position'])

st.markdown("---")

# Upload des fichiers
st.markdown("### 📁 Upload des fichiers")

col1, col2 = st.columns(2)

with col1:
    gsc_old_file = st.file_uploader(
        "**Extraction GSC - Date ancienne**",
        type=['csv'],
        help="Fichier CSV d'extraction GSC par page pour la période ancienne"
    )
    
    gsc_new_file = st.file_uploader(
        "**Extraction GSC - Date actuelle**",
        type=['csv'],
        help="Fichier CSV d'extraction GSC par page pour la période actuelle"
    )

with col2:
    consolidation_file = st.file_uploader(
        "**Consolidation mots-clés GSC**",
        type=['xlsx'],
        help="Fichier Excel de consolidation des mots-clés GSC"
    )
    
    ahrefs_file = st.file_uploader(
        "**Top Pages Ahrefs**",
        type=['csv'],
        help="Export CSV des Top Pages Ahrefs avec comparaison de positions"
    )

st.markdown("---")

def fix_encoding_issues(text):
    """Corrige les problèmes d'encodage courants dans les textes"""
    if not isinstance(text, str):
        return text
    
    # Mapping des séquences mal encodées vers les caractères corrects
    encoding_fixes = {
        '√™': 'ê',   # ê mal encodé
        '√©': 'é',   # é mal encodé
        '√†': 'à',   # à mal encodé
        '√¥': 'ô',   # ô mal encodé
        '√ß': 'ç',   # ç mal encodé
        '√®': 'è',
        '√π': 'ù',
        '√¢': 'â',
        '√Æ': 'î',
        '√¨': 'ì',
        '√´': 'ë',
        '√º': 'ü',
        '√ø': 'ÿ',
        '√î': 'É',
        '√Ä': 'À',
        '√Ö': 'Ç',
        'Ã©': 'é',
        'Ã¨': 'è',
        'Ãª': 'ê',
        'Ã ': 'à',
        'Ã¢': 'â',
        'Ã§': 'ç',
        'Ã´': 'ô',
        'Ã¹': 'ù',
        'Ã®': 'î',
        'Ã¯': 'ï',
        'Ã«': 'ë',
        'Ã¼': 'ü',
        'Ã¿': 'ÿ',
        'Ã‰': 'É',
        'Ã€': 'À',
        'Ã‡': 'Ç',
        'Ã"': 'Ô',
        'Ãˆ': 'È',
        'â€™': "'",
        'â€"': '–',
        'â€"': '—',
        'â€œ': '"',
        'â€': '"',
        'Â ': ' ',
        'Â': '',
    }
    
    for bad, good in encoding_fixes.items():
        text = text.replace(bad, good)
    
    return text

def fix_dataframe_encoding(df):
    """Applique la correction d'encodage à toutes les colonnes texte d'un DataFrame"""
    for col in df.columns:
        if df[col].dtype == 'object':
            df[col] = df[col].apply(lambda x: fix_encoding_issues(x) if isinstance(x, str) else x)
    return df

def read_gsc_csv(file):
    """Lit un fichier CSV GSC avec tentatives multi-encodages (sans chardet)"""
    # Essayer plusieurs encodages courants, mac_roman en priorité pour les fichiers Mac
    encodings_to_try = ['utf-8-sig', 'mac_roman', 'utf-8', 'cp1252', 'latin-1', 'iso-8859-1']
    
    df = None
    for encoding in encodings_to_try:
        try:
            file.seek(0)
            df = pd.read_csv(file, encoding=encoding)
            break
        except (UnicodeDecodeError, UnicodeError, LookupError):
            continue
    
    # Si aucun encodage ne fonctionne, utiliser 'replace' pour les erreurs
    if df is None:
        file.seek(0)
        df = pd.read_csv(file, encoding='utf-8', errors='replace')
    
    # Appliquer la correction d'encodage
    df = fix_dataframe_encoding(df)
    
    return df

def read_consolidation_xlsx(file):
    """Lit le fichier de consolidation Excel"""
    df = pd.read_excel(file)
    df = fix_dataframe_encoding(df)
    return df

def read_ahrefs_csv(file):
    """Lit un fichier CSV Ahrefs (UTF-16, tab-separated)"""
    df = pd.read_csv(file, encoding='utf-16', sep='\t')
    df = fix_dataframe_encoding(df)
    return df

def format_date(date_str, for_sheet_name=False):
    """Formate une date au format xx/xx/xx ou xx-xx-xx pour les noms de feuilles"""
    if pd.isna(date_str):
        return ""
    try:
        date_obj = pd.to_datetime(date_str)
        if for_sheet_name:
            return date_obj.strftime('%d-%m-%y')  # Tirets pour les noms de feuilles
        else:
            return date_obj.strftime('%d/%m/%y')  # Slashes pour les en-têtes
    except:
        return str(date_str)

def calculate_percentage_change(old_val, new_val):
    """Calcule la différence en pourcentage"""
    if pd.isna(old_val) or pd.isna(new_val):
        return None
    
    # Si ancienne valeur = 0, calculer avec base 1 pour éviter division par 0
    if old_val == 0:
        if new_val == 0:
            return 0  # 0 → 0 = 0% de changement
        else:
            # 0 → X = calculer avec base 1
            return ((new_val - 1) / 1) * 100
    
    return ((new_val - old_val) / old_val) * 100

def create_excel_file(df_main, df_gsc_old, df_gsc_new, df_consolidation, df_ahrefs, old_date, new_date, generation_date):
    """Crée le fichier Excel avec mise en forme"""
    wb = Workbook()
    
    # Feuille principale
    ws_main = wb.active
    ws_main.title = "Ré-optimisation"
    
    # Couleurs
    blue_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    green_fill = PatternFill(start_color="548235", end_color="548235", fill_type="solid")
    light_green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    light_red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    white_font = Font(name='Arial', size=10, bold=True, color="FFFFFF")
    normal_font = Font(name='Arial', size=10)
    italic_font = Font(name='Arial', size=10, italic=True)  # Police italique pour "Aucune donnée remontée"
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # En-têtes
    old_date_fmt = format_date(old_date)
    new_date_fmt = format_date(new_date)
    
    headers = [
        "Page",
        "Mots-clés (GSC)",
        "Top Mot-clé (GSC)",
        "Top Mot-clé (Ahrefs)",
        f"Position au {old_date_fmt}",
        f"Position au {new_date_fmt}",
        f"Clics {old_date_fmt}",
        f"Clics {new_date_fmt}",
        "Différence de clics",
        f"Impressions {old_date_fmt}",
        f"Impressions {new_date_fmt}",
        "Différence d'impressions",
        f"CTR {old_date_fmt}",
        f"CTR {new_date_fmt}",
        "Différence de CTR",
        "Position moy.",
        "Briefs de ré-optimisation",
        "Commentaires"
    ]
    
    # Écrire les en-têtes
    for col_idx, header in enumerate(headers, 1):
        cell = ws_main.cell(row=1, column=col_idx, value=header)
        cell.font = white_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border
        # Colonnes G à P (7 à 16) en vert, le reste en bleu
        if 7 <= col_idx <= 16:
            cell.fill = green_fill
        else:
            cell.fill = blue_fill
    
    # Écrire les données (limité à 18 colonnes - A à R)
    max_columns = 18
    for row_idx, row_data in enumerate(df_main.itertuples(index=False), 2):
        for col_idx, value in enumerate(row_data[:max_columns], 1):  # Limiter à 18 colonnes
            # Colonne A : URLs cliquables
            if col_idx == 1:  # Colonne Page (A)
                if pd.notna(value) and value:
                    cell = ws_main.cell(row=row_idx, column=col_idx)
                    cell.value = f'=HYPERLINK("{value}","{value}")'
                    cell.font = Font(name='Arial', size=10, color="0563C1", underline="single")
                else:
                    cell = ws_main.cell(row=row_idx, column=col_idx, value="Aucune donnée remontée")
                    cell.font = italic_font
            # Colonnes texte B, C, D
            elif col_idx in [2, 3, 4]:
                if pd.isna(value) or value == '' or value is None:
                    display_value = "Aucune donnée remontée"
                    cell = ws_main.cell(row=row_idx, column=col_idx, value=display_value)
                    cell.font = italic_font
                else:
                    cell = ws_main.cell(row=row_idx, column=col_idx, value=value)
                    cell.font = normal_font
            # Colonnes numériques E à P (5 à 16)
            elif col_idx in range(5, 17):
                if pd.isna(value) or value is None:
                    display_value = "Aucune donnée remontée"
                    cell = ws_main.cell(row=row_idx, column=col_idx, value=display_value)
                    cell.font = italic_font
                else:
                    cell = ws_main.cell(row=row_idx, column=col_idx, value=value)
                    cell.font = normal_font
            # Colonnes Q et R (17, 18) - laisser vides
            else:
                cell = ws_main.cell(row=row_idx, column=col_idx, value=None)
                cell.font = normal_font
            
            # Alignement : centrer les colonnes numériques (E à P, colonnes 5 à 16)
            if col_idx in range(5, 17):  # Colonnes E à P
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            else:
                cell.alignment = Alignment(vertical='center', wrap_text=True)
            
            cell.border = thin_border
            
            # Format spécifique pour certaines colonnes
            if col_idx in [5, 6, 16]:  # Positions - sans décimales
                if value is not None and not pd.isna(value):
                    cell.number_format = '0'
            elif col_idx in [13, 14]:  # CTR - 1 décimale en pourcentage
                if value is not None and not pd.isna(value):
                    cell.number_format = '0.0%'
            elif col_idx in [9, 12, 15]:  # Différences en pourcentage
                if value is not None and not pd.isna(value):
                    cell.number_format = '0.0%'
    
    # Mise en forme conditionnelle pour les colonnes E et F (positions)
    # Si F < E (amélioration) : E rouge, F vert
    # Si F > E (dégradation) : E vert, F rouge  
    # Si F = E : jaune
    # Ne pas colorer si "Aucune donnée remontée" (texte) - vérifier avec ISNUMBER()
    last_row = len(df_main) + 1
    
    # Colonne E - Position ancienne
    ws_main.conditional_formatting.add(
        f'E2:E{last_row}',
        FormulaRule(formula=['AND(ISNUMBER(E2),ISNUMBER(F2),F2<E2)'], fill=light_red_fill)
    )
    ws_main.conditional_formatting.add(
        f'E2:E{last_row}',
        FormulaRule(formula=['AND(ISNUMBER(E2),ISNUMBER(F2),F2>E2)'], fill=light_green_fill)
    )
    ws_main.conditional_formatting.add(
        f'E2:E{last_row}',
        FormulaRule(formula=['AND(ISNUMBER(E2),ISNUMBER(F2),F2=E2)'], fill=yellow_fill)
    )
    
    # Colonne F - Position actuelle
    ws_main.conditional_formatting.add(
        f'F2:F{last_row}',
        FormulaRule(formula=['AND(ISNUMBER(E2),ISNUMBER(F2),F2<E2)'], fill=light_green_fill)
    )
    ws_main.conditional_formatting.add(
        f'F2:F{last_row}',
        FormulaRule(formula=['AND(ISNUMBER(E2),ISNUMBER(F2),F2>E2)'], fill=light_red_fill)
    )
    ws_main.conditional_formatting.add(
        f'F2:F{last_row}',
        FormulaRule(formula=['AND(ISNUMBER(E2),ISNUMBER(F2),F2=E2)'], fill=yellow_fill)
    )
    
    # Nouvelle mise en forme conditionnelle : gestion des valeurs manquantes avec NOMBRES uniquement
    dark_yellow_fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
    
    # Si donnée numérique en E mais pas en F : E vert, F rouge
    ws_main.conditional_formatting.add(
        f'E2:E{last_row}',
        FormulaRule(formula=['AND(ISNUMBER(E2),NOT(ISNUMBER(F2)))'], fill=light_green_fill)
    )
    ws_main.conditional_formatting.add(
        f'F2:F{last_row}',
        FormulaRule(formula=['AND(ISNUMBER(E2),NOT(ISNUMBER(F2)))'], fill=light_red_fill)
    )
    
    # Si donnée numérique en F mais pas en E : E rouge, F vert
    ws_main.conditional_formatting.add(
        f'E2:E{last_row}',
        FormulaRule(formula=['AND(NOT(ISNUMBER(E2)),ISNUMBER(F2))'], fill=light_red_fill)
    )
    ws_main.conditional_formatting.add(
        f'F2:F{last_row}',
        FormulaRule(formula=['AND(NOT(ISNUMBER(E2)),ISNUMBER(F2))'], fill=light_green_fill)
    )
    
    # Note : Si les deux sont du texte ("Aucune donnée remontée"), pas de mise en forme = fond blanc
    
    # Mise en forme conditionnelle pour I, L, O (différences)
    for col in ['I', 'L', 'O']:
        ws_main.conditional_formatting.add(
            f'{col}2:{col}{last_row}',
            CellIsRule(operator='greaterThan', formula=['0'], fill=light_green_fill)
        )
        ws_main.conditional_formatting.add(
            f'{col}2:{col}{last_row}',
            CellIsRule(operator='lessThan', formula=['0'], fill=light_red_fill)
        )
    
    # Largeurs de colonnes
    column_widths = {
        'A': 60, 'B': 40, 'C': 25, 'D': 25, 'E': 15, 'F': 15,
        'G': 12, 'H': 12, 'I': 15, 'J': 15, 'K': 15, 'L': 18,
        'M': 12, 'N': 12, 'O': 15, 'P': 12, 'Q': 30, 'R': 30
    }
    for col, width in column_widths.items():
        ws_main.column_dimensions[col].width = width
    
    # Hauteur des lignes (68 pixels ≈ 51 points)
    for row in range(1, last_row + 1):
        ws_main.row_dimensions[row].height = 51
    
    # Freeze panes (ligne 1 et colonne A)
    ws_main.freeze_panes = 'B2'
    
    # Supprimer les colonnes S à Z (colonnes 19 à 26) si elles existent
    # On les supprime de droite à gauche pour ne pas décaler les indices
    for col_idx in range(26, 18, -1):  # De Z (26) à S (19)
        col_letter = get_column_letter(col_idx)
        if col_letter in ws_main.column_dimensions:
            ws_main.column_dimensions[col_letter].width = 0
            ws_main.column_dimensions[col_letter].hidden = True
    
    # Feuilles de données brutes
    # GSC ancien
    old_date_sheet = format_date(old_date, for_sheet_name=True)
    ws_gsc_old = wb.create_sheet(title=f"Extraction GSC {old_date_sheet}")
    for r_idx, row in enumerate(dataframe_to_rows(df_gsc_old, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            ws_gsc_old.cell(row=r_idx, column=c_idx, value=value)
    
    # GSC nouveau
    new_date_sheet = format_date(new_date, for_sheet_name=True)
    ws_gsc_new = wb.create_sheet(title=f"Extraction GSC {new_date_sheet}")
    for r_idx, row in enumerate(dataframe_to_rows(df_gsc_new, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            ws_gsc_new.cell(row=r_idx, column=c_idx, value=value)
    
    # Consolidation
    ws_consolidation = wb.create_sheet(title=f"Consolidation GSC {generation_date}")
    for r_idx, row in enumerate(dataframe_to_rows(df_consolidation, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            ws_consolidation.cell(row=r_idx, column=c_idx, value=value)
    
    # Ahrefs
    ws_ahrefs = wb.create_sheet(title=f"Top Pages Ahrefs {generation_date}")
    for r_idx, row in enumerate(dataframe_to_rows(df_ahrefs, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            ws_ahrefs.cell(row=r_idx, column=c_idx, value=value)
    
    return wb

def process_data(df_gsc_old, df_gsc_new, df_consolidation, df_ahrefs, col_names):
    """Traite et fusionne les données"""
    
    # Extraire les dates
    old_date = df_gsc_old[col_names['gsc_start_date']].iloc[0]
    new_date = df_gsc_new[col_names['gsc_end_date']].iloc[0]
    
    # Créer le DataFrame principal basé sur GSC actuel
    df_main = pd.DataFrame()
    df_main['Page'] = df_gsc_new[col_names['gsc_page']]
    
    # Joindre les données GSC ancien
    df_gsc_old_renamed = df_gsc_old[[col_names['gsc_page'], col_names['gsc_clicks'], 
                                     col_names['gsc_impressions'], col_names['gsc_ctr']]].copy()
    df_gsc_old_renamed.columns = ['Page', 'clicks_old', 'impressions_old', 'ctr_old']
    
    df_main = df_main.merge(df_gsc_old_renamed, on='Page', how='left')
    
    # Ajouter les données GSC actuel
    df_main['clicks_new'] = df_gsc_new[col_names['gsc_clicks']].values
    df_main['impressions_new'] = df_gsc_new[col_names['gsc_impressions']].values
    df_main['ctr_new'] = df_gsc_new[col_names['gsc_ctr']].values
    df_main['position_avg'] = df_gsc_new[col_names['gsc_position']].values
    
    # Joindre les mots-clés de la consolidation
    df_consolidation_subset = df_consolidation[[col_names['consolidation_page'], 
                                                 col_names['consolidation_keywords']]].copy()
    df_consolidation_subset.columns = ['Page', 'keywords']
    df_main = df_main.merge(df_consolidation_subset, on='Page', how='left')
    
    # Joindre les données Ahrefs
    df_ahrefs_subset = df_ahrefs[[col_names['ahrefs_url'], col_names['ahrefs_current_top_keyword'],
                                   col_names['ahrefs_previous_position'], col_names['ahrefs_current_position']]].copy()
    df_ahrefs_subset.columns = ['Page', 'top_keyword_ahrefs', 'position_old_ahrefs', 'position_new_ahrefs']
    df_main = df_main.merge(df_ahrefs_subset, on='Page', how='left')
    
    # Calculer les différences
    df_main['clicks_diff'] = df_main.apply(
        lambda row: calculate_percentage_change(row['clicks_old'], row['clicks_new']), axis=1
    )
    df_main['impressions_diff'] = df_main.apply(
        lambda row: calculate_percentage_change(row['impressions_old'], row['impressions_new']), axis=1
    )
    df_main['ctr_diff'] = df_main.apply(
        lambda row: calculate_percentage_change(row['ctr_old'], row['ctr_new']), axis=1
    )
    
    # Extraire le premier mot-clé
    df_main['top_keyword_gsc'] = df_main['keywords'].apply(
        lambda x: str(x).split('\n')[0] if pd.notna(x) else ''
    )
    
    # Convertir les pourcentages de différence en décimal pour Excel
    df_main['clicks_diff'] = df_main['clicks_diff'].apply(lambda x: x/100 if pd.notna(x) else None)
    df_main['impressions_diff'] = df_main['impressions_diff'].apply(lambda x: x/100 if pd.notna(x) else None)
    df_main['ctr_diff'] = df_main['ctr_diff'].apply(lambda x: x/100 if pd.notna(x) else None)
    
    # Réorganiser les colonnes dans l'ordre final
    df_final = pd.DataFrame({
        'Page': df_main['Page'],
        'Mots-clés (GSC)': df_main['keywords'],
        'Top Mot-clé (GSC)': df_main['top_keyword_gsc'],
        'Top Mot-clé (Ahrefs)': df_main['top_keyword_ahrefs'],
        'Position ancienne': df_main['position_old_ahrefs'],
        'Position actuelle': df_main['position_new_ahrefs'],
        'Clics anciens': df_main['clicks_old'],
        'Clics actuels': df_main['clicks_new'],
        'Différence clics': df_main['clicks_diff'],
        'Impressions anciennes': df_main['impressions_old'],
        'Impressions actuelles': df_main['impressions_new'],
        'Différence impressions': df_main['impressions_diff'],
        'CTR ancien': df_main['ctr_old'],
        'CTR actuel': df_main['ctr_new'],
        'Différence CTR': df_main['ctr_diff'],
        'Position moy.': df_main['position_avg'],
        'Briefs': None,  # Laisser vide au lieu de ''
        'Commentaires': None  # Laisser vide au lieu de ''
    })
    
    # S'assurer qu'il n'y a que 18 colonnes (A à R)
    df_final = df_final.iloc[:, :18]
    
    return df_final, old_date, new_date

# Bouton de génération
st.markdown("### 🚀 Générer le fichier")

all_files_uploaded = all([gsc_old_file, gsc_new_file, consolidation_file, ahrefs_file])

if not all_files_uploaded:
    st.warning("⚠️ Veuillez uploader tous les fichiers requis pour générer l'analyse.")

if st.button("Générer le fichier Excel", disabled=not all_files_uploaded, type="primary"):
    
    # Conteneur pour le chronomètre
    timer_placeholder = st.empty()
    progress_placeholder = st.empty()
    
    start_time = time.time()
    
    try:
        # Afficher le chronomètre
        with timer_placeholder.container():
            st.info("⏱️ Génération en cours...")
        
        progress_bar = progress_placeholder.progress(0)
        
        # Lecture des fichiers
        progress_bar.progress(10)
        df_gsc_old = read_gsc_csv(gsc_old_file)
        
        progress_bar.progress(25)
        df_gsc_new = read_gsc_csv(gsc_new_file)
        
        progress_bar.progress(40)
        df_consolidation = read_consolidation_xlsx(consolidation_file)
        
        progress_bar.progress(55)
        df_ahrefs = read_ahrefs_csv(ahrefs_file)
        
        # Traitement des données
        progress_bar.progress(70)
        df_final, old_date, new_date = process_data(
            df_gsc_old, df_gsc_new, df_consolidation, df_ahrefs,
            st.session_state.column_names
        )
        
        # Création du fichier Excel
        progress_bar.progress(85)
        generation_date_fmt = datetime.now().strftime('%d-%m-%y')  # Tirets pour les noms de feuilles
        wb = create_excel_file(
            df_final, df_gsc_old, df_gsc_new, df_consolidation, df_ahrefs,
            old_date, new_date, generation_date_fmt
        )
        
        progress_bar.progress(100)
        
        # Sauvegarder dans un buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        elapsed_time = time.time() - start_time
        
        # Afficher le résultat
        timer_placeholder.empty()
        progress_placeholder.empty()
        
        st.success(f"✅ Fichier généré avec succès en **{elapsed_time:.2f} secondes** !")
        st.info(f"📊 **{len(df_final)}** pages analysées")
        
        # Bouton de téléchargement
        today = datetime.now().strftime('%Y-%m-%d')
        filename = f"reoptimisation_seo_{today}.xlsx"
        
        st.download_button(
            label="📥 Télécharger le fichier Excel",
            data=buffer,
            file_name=filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary"
        )
        
        # Aperçu des données
        with st.expander("👁️ Aperçu des données", expanded=False):
            st.dataframe(df_final.head(20), use_container_width=True)
            
    except Exception as e:
        timer_placeholder.empty()
        progress_placeholder.empty()
        st.error(f"❌ Erreur lors de la génération : {str(e)}")
        st.exception(e)

# Footer
st.markdown("---")
st.markdown(
    """
    <div style='text-align: center; color: #666;'>
        <small>Outil de ré-optimisation SEO | Développé pour l'analyse des performances de contenus</small>
    </div>
    """,
    unsafe_allow_html=True
)
