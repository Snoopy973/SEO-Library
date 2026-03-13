import streamlit as st
import pandas as pd
import re
import io
import os
from pathlib import Path
from datetime import datetime

# ─────────────────────────────────────────────
# CONFIG
# ─────────────────────────────────────────────
if __name__ == "__main__":
    st.set_page_config(
        page_title="🔥 Hotspot Mapper — SEO",
        page_icon="🔥",
        layout="wide",
        initial_sidebar_state="expanded",
    )

# ─────────────────────────────────────────────
# STYLES
# ─────────────────────────────────────────────
st.markdown("""
<style>
    .main .block-container { padding-top: 1rem; }
    .stMetric { background: #1e1e2e; border-radius: 10px; padding: 15px; }
    div[data-testid="stMetricValue"] { font-size: 1.8rem; }

    /* Couleur d'en-tête des tabs */
    .stTabs [data-baseweb="tab"] {
        font-weight: 600;
    }

    /* Forcer les cellules du dataframe à ne pas s'étirer verticalement */
    div[data-testid="stDataFrame"] td {
        max-height: 40px !important;
        overflow: hidden !important;
        text-overflow: ellipsis !important;
        white-space: nowrap !important;
    }
    div[data-testid="stDataFrame"] th {
        white-space: nowrap !important;
    }
</style>
""", unsafe_allow_html=True)




# ─────────────────────────────────────────────
# FONCTIONS UTILITAIRES
# ─────────────────────────────────────────────

def _safe_float(value) -> float:
    """Conversion sécurisée en float (gère virgules françaises)."""
    try:
        return float(str(value).strip().replace(",", ".").replace(" ", ""))
    except (ValueError, TypeError):
        return 0.0


def _safe_int(value) -> int:
    """Conversion sécurisée en int."""
    try:
        return int(_safe_float(value))
    except (ValueError, TypeError):
        return 0


def _filter_brand_keywords(keywords_str: str, brand_regex: str, separator: str = " | ") -> str:
    """
    Supprime les mots-clés qui matchent la regex de marque.
    keywords_str : mots-clés séparés par `separator`
    brand_regex  : pattern regex (case insensitive)
    Retourne la chaîne filtrée.
    """
    if not brand_regex or not keywords_str:
        return keywords_str
    try:
        pattern = re.compile(brand_regex, re.IGNORECASE)
    except re.error:
        return keywords_str  # Regex invalide → on ne filtre rien

    kw_list = [kw.strip() for kw in keywords_str.split(separator)]
    filtered = [kw for kw in kw_list if kw and not pattern.search(kw)]
    return separator.join(filtered)


def parse_gsc_consolidation(df: pd.DataFrame) -> pd.DataFrame:
    """
    Parse le fichier de consolidation GSC.
    Conserve TOUS les mots-clés (séparés par \\n) tels quels pour la colonne finale.
    Rôle : fournir uniquement les mots-clés GSC à associer aux URLs de Screaming Frog.
    """
    # Détection automatique de la colonne URL (peut s'appeler Page, URL, Adresse…)
    url_col = None
    for candidate in ["Page", "URL", "Adresse", "page", "url", "adresse"]:
        if candidate in df.columns:
            url_col = candidate
            break
    if url_col is None:
        raise ValueError(
            f"Colonne URL introuvable dans le fichier GSC. "
            f"Colonnes disponibles : {list(df.columns)}"
        )

    # Détection colonne mots-clés
    kw_col = None
    for candidate in ["Mots clés", "Mots clés GSC", "Keywords", "Query", "Keyword"]:
        if candidate in df.columns:
            kw_col = candidate
            break

    records = []
    for _, row in df.iterrows():
        url = str(row[url_col]).strip() if pd.notna(row[url_col]) else ""

        # Mots-clés : on garde la liste complète en texte multiligne
        all_keywords = ""
        if kw_col and pd.notna(row.get(kw_col)):
            all_keywords = str(row[kw_col]).strip()

        nb_keywords = _safe_int(row.get("Mots clés (#)", 0))
        clicks_total = _safe_int(row.get("Clics (total)", 0))
        impressions_total = _safe_int(row.get("Impressions (total)", 0))

        records.append({
            "URL": url,
            "Meilleurs mot-clés GSC": all_keywords,
            "Nombre Mots-Clés": nb_keywords,
            "Clics Total GSC": clicks_total,
            "Impressions Total GSC": impressions_total,
        })

    return pd.DataFrame(records)


def parse_screaming_frog(df: pd.DataFrame) -> pd.DataFrame:
    """
    Parse le fichier internal_html de Screaming Frog.
    Extrait : URL, Title, Longueur Title, H1, Longueur H1,
              Meta Description, Longueur Meta Description, Code HTTP, Indexabilité.
    (Clics et Impressions viennent du fichier GSC pour la cohérence des données.)
    """
    col_map = {
        "Adresse": "URL",
        "Title 1": "Title 1",
        "Longueur du Title 1": "Longueur du Title 1",
        "H1-1": "H1-1",
        "Longueur du H1-1": "Longueur du H1-1",
        "Meta Description 1": "Meta Description 1",
        "Longueur de la Meta Description 1": "Longueur de la Meta Description 1",
        "Code HTTP": "Code HTTP",
        "Indexabilité": "Indexabilité",
        "Language": "Language",
    }

    available_cols = {k: v for k, v in col_map.items() if k in df.columns}
    result = df[list(available_cols.keys())].rename(columns=available_cols).copy()

    # Nettoyage URL
    if "URL" in result.columns:
        result["URL"] = result["URL"].astype(str).str.strip()

    # Conversion numérique (virgules françaises)
    for col in ["Longueur du Title 1", "Longueur du H1-1", "Longueur de la Meta Description 1"]:
        if col in result.columns:
            result[col] = result[col].apply(_safe_float)

    if "Code HTTP" in result.columns:
        result["Code HTTP"] = result["Code HTTP"].apply(_safe_int)

    return result


def merge_and_build(df_sf: pd.DataFrame, df_gsc: pd.DataFrame, sheet_type: str = "title", brand_regex: str = "") -> pd.DataFrame:
    """
    Fusionne SF + GSC sur l'URL et construit le tableau final.
    - SF fournit : URL, Title/H1/Meta Description (selon sheet_type), Code HTTP, Indexabilité
    - GSC fournit : Clics, Impressions, Mots-clés

    sheet_type: "title", "h1" ou "meta_description"
    brand_regex: regex pour exclure les mots-clés de marque
    """
    # LEFT JOIN : SF est la base, GSC enrichit
    merged = pd.merge(df_sf, df_gsc, on="URL", how="left")

    # Remplir les NaN des colonnes GSC pour les URLs sans match
    for col in ["Meilleurs mot-clés GSC"]:
        if col in merged.columns:
            merged[col] = merged[col].fillna("")
    for col in ["Nombre Mots-Clés", "Clics Total GSC", "Impressions Total GSC"]:
        if col in merged.columns:
            merged[col] = merged[col].fillna(0)

    # Clics et Impressions = depuis GSC (cohérence avec les mots-clés)
    merged["Clics"] = merged["Clics Total GSC"]
    merged["Impressions"] = merged["Impressions Total GSC"]

    # Nettoyer les \n dans les mots-clés pour l'affichage Streamlit + CSV
    if "Meilleurs mot-clés GSC" in merged.columns:
        merged["Meilleurs mot-clés GSC"] = (
            merged["Meilleurs mot-clés GSC"]
            .astype(str)
            .str.replace("\n", " | ", regex=False)
            .str.replace("nan", "", regex=False)
            .str.strip()
            .str.strip("|")
            .str.strip()
        )

        # Filtrer les mots-clés de marque si regex fournie
        if brand_regex:
            merged["Meilleurs mot-clés GSC"] = merged["Meilleurs mot-clés GSC"].apply(
                lambda kw: _filter_brand_keywords(kw, brand_regex, separator=" | ")
            )

    # CTR calculé
    merged["CTR"] = merged.apply(
        lambda r: round(r["Clics"] / r["Impressions"], 5)
        if r["Impressions"] > 0 else 0,
        axis=1
    )

    # Colonnes vides (à remplir manuellement ou en phase 2)
    merged["Type"] = ""
    if "Language" not in merged.columns:
        merged["Language"] = ""
    else:
        merged["Language"] = merged["Language"].fillna("")
    merged["Positions"] = ""                    # Phase 2 — DataForSEO
    # Pré-remplir "Mot-clé sélectionné" avec le 1er mot-clé GSC
    if "Meilleurs mot-clés GSC" in merged.columns:
        merged["Mot-clé sélectionné"] = (
            merged["Meilleurs mot-clés GSC"]
            .astype(str)
            .str.split(r"\s*\|\s*", regex=True)
            .str[0]
            .fillna("")
        )
    else:
        merged["Mot-clé sélectionné"] = ""
    merged["Commentaire"] = ""

    # ── Colonnes et ordre selon le type de feuille ──
    if sheet_type == "title":
        merged["Balise title (proposée)"] = ""
        merged["Title length (proposé)"] = ""
        final_cols = [
            "URL", "Type", "Language", "Clics", "Impressions", "CTR",
            "Positions", "Mot-clé sélectionné", "Meilleurs mot-clés GSC",
            "Title 1", "Longueur du Title 1",
            "Balise title (proposée)", "Title length (proposé)",
            "Commentaire",
            "Indexabilité", "Code HTTP", "Nombre Mots-Clés",
        ]
    elif sheet_type == "h1":
        merged["H1 (proposé)"] = ""
        merged["H1 length (proposé)"] = ""
        final_cols = [
            "URL", "Type", "Language", "Clics", "Impressions", "CTR",
            "Positions", "Mot-clé sélectionné", "Meilleurs mot-clés GSC",
            "H1-1", "Longueur du H1-1",
            "H1 (proposé)", "H1 length (proposé)",
            "Commentaire",
            "Indexabilité", "Code HTTP", "Nombre Mots-Clés",
        ]
    elif sheet_type == "meta_description":
        merged["Meta description (proposée)"] = ""
        merged["Meta desc. length (proposé)"] = ""
        final_cols = [
            "URL", "Type", "Language", "Clics", "Impressions", "CTR",
            "Positions", "Mot-clé sélectionné", "Meilleurs mot-clés GSC",
            "Meta Description 1", "Longueur de la Meta Description 1",
            "Meta description (proposée)", "Meta desc. length (proposé)",
            "Commentaire",
            "Indexabilité", "Code HTTP", "Nombre Mots-Clés",
        ]

    final_cols = [c for c in final_cols if c in merged.columns]
    return merged[final_cols]


def detect_hotspots(df: pd.DataFrame) -> pd.DataFrame:
    """
    Détecte les points chauds : pages à fort potentiel d'optimisation title/CTR.
    Score de chaleur = impressions normalisées × (1 - CTR)
    """
    df = df.copy()

    if "Impressions" in df.columns and "CTR" in df.columns:
        max_imp = df["Impressions"].max() if df["Impressions"].max() > 0 else 1
        df["Score Chaleur"] = (
            (df["Impressions"] / max_imp) * 100 *
            (1 - df["CTR"])
        ).round(1)

        df["Statut"] = df.apply(_categorize, axis=1)
    else:
        df["Score Chaleur"] = 0
        df["Statut"] = "—"

    return df


def _categorize(row) -> str:
    """Catégorise une page selon son potentiel."""
    imp = row.get("Impressions", 0)
    ctr = row.get("CTR", 0)
    clics = row.get("Clics", 0)
    has_kw = bool(str(row.get("Meilleurs mot-clés GSC", "")).strip())

    # Pas de données exploitables → pas un point chaud
    if imp == 0 or (clics == 0 and imp < 50):
        return "✅ OK"

    if imp >= 500 and ctr < 0.02:
        return "🔥 Chaud"
    elif imp >= 100 and ctr < 0.03:
        return "🟠 Tiède"
    elif imp >= 20 and ctr < 0.05:
        return "🟡 À surveiller"
    else:
        return "✅ OK"


def _style_excel_sheet(ws, df):
    """Applique le style MDPC (header bleu foncé, alternance, formats) à une feuille Excel."""
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

    # ── Couleurs du thème (calquées sur le modèle MDPC) ──
    HEADER_FILL = PatternFill(start_color="1B2A4A", end_color="1B2A4A", fill_type="solid")
    HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=9)
    BODY_FONT = Font(name="Arial", size=9, color="333333")
    BORDER_THIN = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )
    CENTER = Alignment(horizontal="center", vertical="center", wrap_text=False)
    CENTER_WRAP = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LEFT_WRAP = Alignment(horizontal="left", vertical="center", wrap_text=True)

    ROW_EVEN = PatternFill(start_color="F2F6FC", end_color="F2F6FC", fill_type="solid")
    ROW_ODD = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    # Couleurs par statut
    STATUT_FILLS = {
        "🔥 Chaud":       PatternFill(start_color="FFE0E0", end_color="FFE0E0", fill_type="solid"),
        "🟠 Tiède":       PatternFill(start_color="FFF0D0", end_color="FFF0D0", fill_type="solid"),
        "🟡 À surveiller": PatternFill(start_color="FFFDE0", end_color="FFFDE0", fill_type="solid"),
        "✅ OK":           PatternFill(start_color="E0FFE0", end_color="E0FFE0", fill_type="solid"),
    }

    # Colonnes entières (format nombre sans décimale)
    int_cols = {
        "Clics", "Impressions",
        "Longueur du Title 1", "Longueur du H1-1", "Longueur de la Meta Description 1",
    }
    # Colonnes texte large (alignement gauche)
    text_wide_cols = {
        "URL",
        "Title 1", "Balise title (proposée)",
        "H1-1", "H1 (proposé)",
        "Meta Description 1", "Meta description (proposée)",
    }

    # ── Style en-tête (ligne 1) ──
    for col_idx in range(1, len(df.columns) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER_WRAP
        cell.border = BORDER_THIN

    # Trouver l'index de la colonne "Statut" (si elle existe)
    statut_col_idx = None
    for ci, cn in enumerate(df.columns):
        if cn == "Statut":
            statut_col_idx = ci
            break

    # ── Style corps (lignes 2+) ──
    for row_idx in range(2, len(df) + 2):
        # Couleur basée sur le statut si disponible
        default_fill = ROW_EVEN if row_idx % 2 == 0 else ROW_ODD
        row_fill = default_fill

        if statut_col_idx is not None:
            statut_val = str(df.iloc[row_idx - 2, statut_col_idx])
            for statut_key, statut_fill in STATUT_FILLS.items():
                if statut_key in statut_val:
                    row_fill = statut_fill
                    break

        for col_idx in range(1, len(df.columns) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            col_name = df.columns[col_idx - 1]

            cell.fill = row_fill
            cell.font = BODY_FONT
            cell.border = BORDER_THIN

            # Alignement
            if col_name == "Meilleurs mot-clés GSC":
                cell.alignment = Alignment(
                    horizontal="left", vertical="center", wrap_text=False
                )
            elif col_name in text_wide_cols:
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
            else:
                cell.alignment = CENTER

            # Format nombres entiers
            if col_name in int_cols:
                cell.number_format = '#,##0'
                try:
                    val = cell.value
                    if val is not None and val != "":
                        cell.value = int(float(val))
                except (ValueError, TypeError):
                    pass

            # Format CTR en pourcentage
            if col_name == "CTR":
                cell.number_format = '0,00%'

    # ── Largeur des colonnes ──
    col_widths = {
        "URL": 55,
        "Type": 12,
        "Language": 8,
        "Clics": 10,
        "Impressions": 14,
        "CTR": 10,
        "Positions": 12,
        "Mot-clé sélectionné": 25,
        "Meilleurs mot-clés GSC": 40,
        "Title 1": 50,
        "Longueur du Title 1": 12,
        "Balise title (proposée)": 50,
        "Title length (proposé)": 12,
        "H1-1": 50,
        "Longueur du H1-1": 12,
        "H1 (proposé)": 50,
        "H1 length (proposé)": 12,
        "Meta Description 1": 60,
        "Longueur de la Meta Description 1": 14,
        "Meta description (proposée)": 60,
        "Meta desc. length (proposé)": 14,
        "Commentaire": 30,
        "Indexabilité": 14,
        "Code HTTP": 10,
        "Nombre Mots-Clés": 12,
    }
    for col_idx, col_name in enumerate(df.columns, 1):
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        ws.column_dimensions[col_letter].width = col_widths.get(col_name, 15)

    # ── Figer la première ligne + première colonne ──
    ws.freeze_panes = "B2"
    ws.row_dimensions[1].height = 22

    # ── Hauteur fixe des lignes du corps (empêche l'étirement) ──
    for row_idx in range(2, len(df) + 2):
        ws.row_dimensions[row_idx].height = 15

    # ── Filtre automatique ──
    last_col_letter = ws.cell(row=1, column=len(df.columns)).column_letter
    ws.auto_filter.ref = f"A1:{last_col_letter}{len(df) + 1}"


def to_excel_download(df: pd.DataFrame) -> bytes:
    """Convertit un seul DataFrame en fichier Excel stylisé (1 feuille)."""
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="MDPC Title")
        _style_excel_sheet(writer.sheets["MDPC Title"], df)
    return output.getvalue()


def to_excel_multi_sheets(
    df_title: pd.DataFrame,
    df_h1: pd.DataFrame,
    df_meta: pd.DataFrame,
) -> bytes:
    """
    Génère un fichier Excel avec 3 feuilles stylisées :
    - MDPC Title
    - MDPC H1
    - MDPC Meta Description
    """
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        # ── Feuille 1 : MDPC Title ──
        df_title.to_excel(writer, index=False, sheet_name="MDPC Title")
        _style_excel_sheet(writer.sheets["MDPC Title"], df_title)

        # ── Feuille 2 : MDPC H1 ──
        df_h1.to_excel(writer, index=False, sheet_name="MDPC H1")
        _style_excel_sheet(writer.sheets["MDPC H1"], df_h1)

        # ── Feuille 3 : MDPC Meta Description ──
        df_meta.to_excel(writer, index=False, sheet_name="MDPC Meta Description")
        _style_excel_sheet(writer.sheets["MDPC Meta Description"], df_meta)

    return output.getvalue()


# ─────────────────────────────────────────────
# EXPORT GOOGLE SHEETS
# ─────────────────────────────────────────────

def _get_credentials_path():
    """Cherche le fichier credentials.json dans le dossier du projet."""
    script_dir = Path(__file__).parent
    cred_path = script_dir / "credentials.json"
    if cred_path.exists():
        return str(cred_path)
    return None


def _get_gspread_client():
    """Retourne un client gspread authentifié avec le compte de service."""
    import gspread
    from google.oauth2.service_account import Credentials

    cred_path = _get_credentials_path()
    if not cred_path:
        raise FileNotFoundError(
            "Fichier credentials.json introuvable dans le dossier du projet."
        )

    scopes = [
        "https://www.googleapis.com/auth/spreadsheets",
        "https://www.googleapis.com/auth/drive",
    ]
    creds = Credentials.from_service_account_file(cred_path, scopes=scopes)
    return gspread.authorize(creds)


def _populate_spreadsheet(sh, df_title, df_h1, df_meta):
    """Remplit un Google Sheet avec les 3 feuilles MDPC + Config."""
    import gspread

    sheets_data = {
        "MDPC Title": df_title,
        "MDPC H1": df_h1,
        "MDPC Meta Description": df_meta,
    }

    for sheet_name, df in sheets_data.items():
        # ── Supprimer la feuille existante pour repartir à zéro ──
        try:
            old_ws = sh.worksheet(sheet_name)
            sh.del_worksheet(old_ws)
        except gspread.exceptions.WorksheetNotFound:
            pass

        # ── Créer une feuille neuve ──
        ws = sh.add_worksheet(title=sheet_name, rows=len(df) + 1, cols=len(df.columns))

        # ── Préparer les données ──
        headers = df.columns.tolist()
        rows = []
        for _, row in df.iterrows():
            row_data = []
            for col in df.columns:
                val = row[col]
                if pd.isna(val):
                    row_data.append("")
                elif isinstance(val, float):
                    row_data.append(val)
                elif isinstance(val, (int,)):
                    row_data.append(int(val))
                else:
                    row_data.append(str(val))
            rows.append(row_data)

        # ── Écriture en batch ──
        all_data = [headers] + rows
        ws.update(all_data, value_input_option="RAW")

        # ── Mise en forme via API batch ──
        _format_gsheet_worksheet(sh, ws, df)

    # ── Supprimer la feuille par défaut "Sheet1" / "Feuille 1" si elle existe ──
    for default_name in ["Sheet1", "Feuille 1", "Feuille1"]:
        try:
            default_ws = sh.worksheet(default_name)
            sh.del_worksheet(default_ws)
        except Exception:
            pass

    # ── Créer l'onglet Config avec paramètres pré-remplis ──
    _create_config_gsheet(sh)


PERSONAL_SHEET_URL = "https://docs.google.com/spreadsheets/d/1GOeHrdejlusHTsu9mQ00-rwpsNKS-rvD98Q1637zJNg/edit"


def _update_personal_sheet(
    df_title: pd.DataFrame,
    df_h1: pd.DataFrame,
    df_meta: pd.DataFrame,
):
    """
    Met à jour le Google Sheet personnel existant.
    Retourne l'URL du Sheet.
    """
    gc = _get_gspread_client()
    sh = gc.open_by_url(PERSONAL_SHEET_URL)
    _populate_spreadsheet(sh, df_title, df_h1, df_meta)
    return sh.url


def _create_config_gsheet(sh):
    """Crée (ou recrée) l'onglet Config avec les 4 paramètres DataForSEO + descriptions."""
    import gspread

    # Supprimer l'ancien onglet Config s'il existe
    try:
        old = sh.worksheet("Config")
        sh.del_worksheet(old)
    except gspread.exceptions.WorksheetNotFound:
        pass

    ws = sh.add_worksheet(title="Config", rows=7, cols=3)

    # Données : header + 4 paramètres
    data = [
        ["Paramètre", "Valeur", "Description"],
        [
            "pays",
            "fr",
            "Code pays ISO pour la SERP DataForSEO — ex : fr (France), us (USA), "
            "gb (Royaume-Uni), de (Allemagne), es (Espagne), it (Italie), "
            "be (Belgique), ch (Suisse), ca (Canada), br (Brésil), nl (Pays-Bas), pt (Portugal)",
        ],
        [
            "langue_serp",
            "fr",
            "Code langue ISO pour les résultats — ex : fr (français), en (anglais), "
            "de (allemand), es (espagnol), it (italien), nl (néerlandais), pt (portugais). "
            "Doit correspondre au pays choisi (ex : pays=fr → langue=fr, pays=us → langue=en)",
        ],
        [
            "device",
            "desktop",
            "Type d'appareil pour la SERP — valeurs possibles : desktop ou mobile. "
            "Les positions peuvent varier entre desktop et mobile.",
        ],
        [
            "domaine_cible",
            "",
            "Votre domaine SANS www ni https — ex : ranchcomputing.com. "
            "Sert à repérer vos URLs dans les résultats de recherche DataForSEO.",
        ],
    ]

    ws.update(data, value_input_option="RAW")

    sheet_id = ws.id
    requests = []

    # Header : fond bleu foncé, texte blanc, gras
    requests.append({
        "repeatCell": {
            "range": {"sheetId": sheet_id, "startRowIndex": 0, "endRowIndex": 1,
                       "startColumnIndex": 0, "endColumnIndex": 3},
            "cell": {
                "userEnteredFormat": {
                    "backgroundColor": {"red": 0.106, "green": 0.165, "blue": 0.29},
                    "textFormat": {"foregroundColor": {"red": 1, "green": 1, "blue": 1},
                                   "bold": True, "fontFamily": "Arial", "fontSize": 10},
                    "horizontalAlignment": "CENTER",
                    "verticalAlignment": "MIDDLE",
                }
            },
            "fields": "userEnteredFormat",
        }
    })

    # Colonne Description : gris, italique
    requests.append({
        "repeatCell": {
            "range": {"sheetId": sheet_id, "startRowIndex": 1, "endRowIndex": 6,
                       "startColumnIndex": 2, "endColumnIndex": 3},
            "cell": {
                "userEnteredFormat": {
                    "textFormat": {"foregroundColor": {"red": 0.5, "green": 0.5, "blue": 0.5},
                                   "fontFamily": "Arial", "fontSize": 9, "italic": True},
                    "wrapStrategy": "WRAP",
                }
            },
            "fields": "userEnteredFormat(textFormat,wrapStrategy)",
        }
    })

    # Colonne Paramètre : gras
    requests.append({
        "repeatCell": {
            "range": {"sheetId": sheet_id, "startRowIndex": 1, "endRowIndex": 6,
                       "startColumnIndex": 0, "endColumnIndex": 1},
            "cell": {
                "userEnteredFormat": {
                    "textFormat": {"bold": True, "fontFamily": "Arial", "fontSize": 10},
                }
            },
            "fields": "userEnteredFormat(textFormat)",
        }
    })

    # Colonne Valeur : fond jaune clair pour indiquer que c'est éditable
    requests.append({
        "repeatCell": {
            "range": {"sheetId": sheet_id, "startRowIndex": 1, "endRowIndex": 6,
                       "startColumnIndex": 1, "endColumnIndex": 2},
            "cell": {
                "userEnteredFormat": {
                    "backgroundColor": {"red": 1, "green": 0.98, "blue": 0.8},
                    "textFormat": {"fontFamily": "Arial", "fontSize": 10},
                }
            },
            "fields": "userEnteredFormat(backgroundColor,textFormat)",
        }
    })

    # Largeurs des colonnes
    for ci, w in enumerate([180, 220, 550]):
        requests.append({
            "updateDimensionProperties": {
                "range": {"sheetId": sheet_id, "dimension": "COLUMNS",
                           "startIndex": ci, "endIndex": ci + 1},
                "properties": {"pixelSize": w},
                "fields": "pixelSize",
            }
        })

    # Hauteur header (35px)
    requests.append({
        "updateDimensionProperties": {
            "range": {"sheetId": sheet_id, "dimension": "ROWS",
                       "startIndex": 0, "endIndex": 1},
            "properties": {"pixelSize": 35},
            "fields": "pixelSize",
        }
    })

    # Hauteur lignes du corps (40px pour laisser de la place aux descriptions)
    requests.append({
        "updateDimensionProperties": {
            "range": {"sheetId": sheet_id, "dimension": "ROWS",
                       "startIndex": 1, "endIndex": 6},
            "properties": {"pixelSize": 40},
            "fields": "pixelSize",
        }
    })

    sh.batch_update({"requests": requests})


def _format_gsheet_worksheet(sh, ws, df):
    """Applique la mise en forme MDPC à une feuille Google Sheets via l'API batch."""
    import gspread

    n_rows = len(df) + 1  # +1 pour le header
    n_cols = len(df.columns)
    headers = df.columns.tolist()
    sheet_id = ws.id

    # Index des colonnes
    ctr_idx = headers.index("CTR") if "CTR" in headers else -1
    url_idx = headers.index("URL") if "URL" in headers else -1

    # ── Requêtes batch ──
    requests = []

    # 1. En-tête : fond bleu foncé, texte blanc, gras
    requests.append({
        "repeatCell": {
            "range": {"sheetId": sheet_id, "startRowIndex": 0, "endRowIndex": 1,
                       "startColumnIndex": 0, "endColumnIndex": n_cols},
            "cell": {
                "userEnteredFormat": {
                    "backgroundColor": {"red": 0.106, "green": 0.165, "blue": 0.29},
                    "textFormat": {"foregroundColor": {"red": 1, "green": 1, "blue": 1},
                                   "bold": True, "fontFamily": "Arial", "fontSize": 10},
                    "horizontalAlignment": "CENTER",
                    "verticalAlignment": "MIDDLE",
                }
            },
            "fields": "userEnteredFormat",
        }
    })

    # 2. Corps : police Arial 10, pas de wrap, clip
    requests.append({
        "repeatCell": {
            "range": {"sheetId": sheet_id, "startRowIndex": 1, "endRowIndex": n_rows,
                       "startColumnIndex": 0, "endColumnIndex": n_cols},
            "cell": {
                "userEnteredFormat": {
                    "textFormat": {"fontFamily": "Arial", "fontSize": 10,
                                   "foregroundColor": {"red": 0.2, "green": 0.2, "blue": 0.2}},
                    "verticalAlignment": "MIDDLE",
                    "wrapStrategy": "CLIP",
                }
            },
            "fields": "userEnteredFormat(textFormat,verticalAlignment,wrapStrategy)",
        }
    })

    # 3. Format CTR en pourcentage
    if ctr_idx >= 0:
        requests.append({
            "repeatCell": {
                "range": {"sheetId": sheet_id, "startRowIndex": 1, "endRowIndex": n_rows,
                           "startColumnIndex": ctr_idx, "endColumnIndex": ctr_idx + 1},
                "cell": {
                    "userEnteredFormat": {
                        "numberFormat": {"type": "PERCENT", "pattern": "0.00%"},
                        "horizontalAlignment": "CENTER",
                    }
                },
                "fields": "userEnteredFormat(numberFormat,horizontalAlignment)",
            }
        })

    # 4. Format nombres entiers (Clics, Impressions, Longueurs)
    int_cols = {"Clics", "Impressions", "Longueur du Title 1", "Longueur du H1-1",
                "Longueur de la Meta Description 1"}
    for col_name in int_cols:
        if col_name in headers:
            ci = headers.index(col_name)
            requests.append({
                "repeatCell": {
                    "range": {"sheetId": sheet_id, "startRowIndex": 1, "endRowIndex": n_rows,
                               "startColumnIndex": ci, "endColumnIndex": ci + 1},
                    "cell": {
                        "userEnteredFormat": {
                            "numberFormat": {"type": "NUMBER", "pattern": "#,##0"},
                            "horizontalAlignment": "CENTER",
                        }
                    },
                    "fields": "userEnteredFormat(numberFormat,horizontalAlignment)",
                }
            })

    # 5. Freeze header + colonne A
    requests.append({
        "updateSheetProperties": {
            "properties": {
                "sheetId": sheet_id,
                "gridProperties": {"frozenRowCount": 1, "frozenColumnCount": 1},
            },
            "fields": "gridProperties.frozenRowCount,gridProperties.frozenColumnCount",
        }
    })

    # 6. Largeurs des colonnes
    col_widths = {
        "URL": 400, "Type": 80, "Language": 60, "Clics": 70, "Impressions": 100,
        "CTR": 70, "Positions": 80, "Mot-clé sélectionné": 180,
        "Meilleurs mot-clés GSC": 280, "Title 1": 350, "H1-1": 350,
        "Meta Description 1": 400, "Longueur du Title 1": 80,
        "Longueur du H1-1": 80, "Longueur de la Meta Description 1": 90,
        "Balise title (proposée)": 350, "Title length (proposé)": 80,
        "H1 (proposé)": 350, "H1 length (proposé)": 80,
        "Meta description (proposée)": 400, "Meta desc. length (proposé)": 90,
        "Commentaire": 200,
    }
    for ci, col_name in enumerate(headers):
        w = col_widths.get(col_name, 100)
        requests.append({
            "updateDimensionProperties": {
                "range": {"sheetId": sheet_id, "dimension": "COLUMNS",
                           "startIndex": ci, "endIndex": ci + 1},
                "properties": {"pixelSize": w},
                "fields": "pixelSize",
            }
        })

    # 7. Hauteur fixe des lignes du corps (21px)
    requests.append({
        "updateDimensionProperties": {
            "range": {"sheetId": sheet_id, "dimension": "ROWS",
                       "startIndex": 1, "endIndex": n_rows},
            "properties": {"pixelSize": 21},
            "fields": "pixelSize",
        }
    })

    # 8. Hauteur header (35px)
    requests.append({
        "updateDimensionProperties": {
            "range": {"sheetId": sheet_id, "dimension": "ROWS",
                       "startIndex": 0, "endIndex": 1},
            "properties": {"pixelSize": 35},
            "fields": "pixelSize",
        }
    })

    # 9. Filtre automatique
    requests.append({
        "setBasicFilter": {
            "filter": {
                "range": {"sheetId": sheet_id, "startRowIndex": 0, "endRowIndex": n_rows,
                           "startColumnIndex": 0, "endColumnIndex": n_cols},
            }
        }
    })

    # ── Exécuter toutes les requêtes en une fois ──
    sh.batch_update({"requests": requests})


# ─────────────────────────────────────────────
# COLONNES AFFICHÉES DANS STREAMLIT
# ─────────────────────────────────────────────

COLUMN_CONFIG = {
    "Statut": st.column_config.TextColumn("Statut", width="small"),
    "URL": st.column_config.LinkColumn("URL", width="medium"),
    "Type": st.column_config.TextColumn("Type", width="small"),
    "Language": st.column_config.TextColumn("Language", width="small"),
    "CTR": st.column_config.NumberColumn("CTR (%)", format="%.2f %%", width="small"),
    "Clics": st.column_config.NumberColumn("Clics", format="%d", width="small"),
    "Impressions": st.column_config.NumberColumn("Impr.", format="%d", width="small"),
    "Mot-clé sélectionné": st.column_config.TextColumn("Mot-clé sél.", width="small"),
    "Meilleurs mot-clés GSC": st.column_config.TextColumn("Top mots-clés GSC", width="medium"),
    "Title 1": st.column_config.TextColumn("Title", width="medium"),
    "H1-1": st.column_config.TextColumn("H1", width="medium"),
    "Meta Description 1": st.column_config.TextColumn("Meta Desc.", width="medium"),
    "Longueur du Title 1": st.column_config.NumberColumn("Long.", format="%d", width="small"),
    "Longueur du H1-1": st.column_config.NumberColumn("Long.", format="%d", width="small"),
    "Longueur de la Meta Description 1": st.column_config.NumberColumn("Long.", format="%d", width="small"),
    "Score Chaleur": st.column_config.ProgressColumn(
        "🔥 Score", min_value=0, max_value=100, format="%.0f",
    ),
    "Commentaire": st.column_config.TextColumn("Commentaire", width="small"),
}

# Mapping des types de feuilles
SHEET_TYPES = {
    "title": {
        "label": "🏷️ MDPC Title",
        "tag_col": "Title 1",
        "length_col": "Longueur du Title 1",
        "proposed_col": "Balise title (proposée)",
        "proposed_len_col": "Title length (proposé)",
    },
    "h1": {
        "label": "📝 MDPC H1",
        "tag_col": "H1-1",
        "length_col": "Longueur du H1-1",
        "proposed_col": "H1 (proposé)",
        "proposed_len_col": "H1 length (proposé)",
    },
    "meta_description": {
        "label": "📋 MDPC Meta Description",
        "tag_col": "Meta Description 1",
        "length_col": "Longueur de la Meta Description 1",
        "proposed_col": "Meta description (proposée)",
        "proposed_len_col": "Meta desc. length (proposé)",
    },
}


# ─────────────────────────────────────────────
# INTERFACE STREAMLIT
# ─────────────────────────────────────────────

def _apply_filters(df, filter_lang, filter_indexable, filter_http_200, filter_min_impressions, filter_min_clicks):
    """Applique les filtres communs à un DataFrame."""
    if filter_lang:
        df = df[df["URL"].apply(lambda u: any(lang in str(u) for lang in filter_lang))]
    if filter_indexable and "Indexabilité" in df.columns:
        df = df[df["Indexabilité"] == "Indexable"]
    if filter_http_200 and "Code HTTP" in df.columns:
        df = df[df["Code HTTP"] == 200]
    if filter_min_impressions > 0 and "Impressions" in df.columns:
        df = df[df["Impressions"] >= filter_min_impressions]
    if filter_min_clicks > 0 and "Clics" in df.columns:
        df = df[df["Clics"] >= filter_min_clicks]
    return df


def _prepare_excel_df(df):
    """Prépare un DataFrame pour l'export Excel : reconvertit ' | ' en \\n pour les mots-clés."""
    excel_df = df.copy()
    if "Meilleurs mot-clés GSC" in excel_df.columns:
        excel_df["Meilleurs mot-clés GSC"] = (
            excel_df["Meilleurs mot-clés GSC"]
            .astype(str)
            .str.replace(" | ", "\n", regex=False)
        )
    return excel_df


def _get_export_cols(sheet_type):
    """Retourne la liste des colonnes d'export selon le type de feuille."""
    info = SHEET_TYPES[sheet_type]
    return [
        "URL", "Type", "Language", "Clics", "Impressions", "CTR",
        "Positions", "Mot-clé sélectionné", "Meilleurs mot-clés GSC",
        info["tag_col"], info["length_col"],
        info["proposed_col"], info["proposed_len_col"],
        "Commentaire",
    ]


def _get_display_cols(sheet_type):
    """Retourne la liste des colonnes d'affichage Streamlit (sans colonnes vides phase 2)."""
    info = SHEET_TYPES[sheet_type]
    return [
        "Statut", "URL", "Type", "Language",
        "Clics", "Impressions", "CTR",
        "Mot-clé sélectionné", "Meilleurs mot-clés GSC",
        info["tag_col"], info["length_col"],
        "Commentaire",
    ]


def _render_sheet_tab(result, sheet_type, tab_key_suffix=""):
    """
    Affiche le contenu d'un onglet de feuille dans Streamlit :
    métriques, points chauds, tableau complet, stats, recherche.
    """
    info = SHEET_TYPES[sheet_type]
    display_cols = [c for c in _get_display_cols(sheet_type) if c in result.columns]
    search_cols = ["URL", "Meilleurs mot-clés GSC", info["tag_col"]]

    # Convertir CTR décimal → pourcentage pour l'affichage Streamlit
    result = result.copy()
    if "CTR" in result.columns:
        result["CTR"] = (result["CTR"] * 100).round(2)

    # Tronquer les mots-clés GSC à max 3 pour l'affichage Streamlit (compact)
    if "Meilleurs mot-clés GSC" in result.columns:
        def _truncate_kw(kw_str, max_kw=3):
            if not kw_str or kw_str == "":
                return ""
            parts = [p.strip() for p in str(kw_str).split(" | ") if p.strip()]
            if len(parts) <= max_kw:
                return ", ".join(parts)
            return ", ".join(parts[:max_kw]) + f" (+{len(parts) - max_kw})"
        result["Meilleurs mot-clés GSC"] = result["Meilleurs mot-clés GSC"].apply(_truncate_kw)

    # ── Métriques ──
    urls_with_kw = len(result[result["Meilleurs mot-clés GSC"].astype(str).str.len() > 0]) if "Meilleurs mot-clés GSC" in result.columns else 0
    urls_no_kw = len(result) - urls_with_kw
    if urls_no_kw > 0:
        st.caption(f"🔗 {urls_with_kw} URLs matchées avec GSC — {urls_no_kw} sans mots-clés GSC")

    c1, c2, c3, c4, c5, c6 = st.columns(6)
    n_chaud = len(result[result["Statut"] == "🔥 Chaud"])
    n_tiede = len(result[result["Statut"] == "🟠 Tiède"])
    n_surveiller = len(result[result["Statut"] == "🟡 À surveiller"])
    with c1:
        st.metric("📄 URLs", len(result))
    with c2:
        st.markdown(
            f'<div style="background:rgba(255,75,75,0.2); border-left:4px solid #FF4B4B; border-radius:8px; padding:12px;">'
            f'<span style="font-size:0.85rem; color:#aaa;">🔥 Chaud</span><br>'
            f'<span style="font-size:1.8rem; font-weight:700; color:#FF4B4B;">{n_chaud}</span></div>',
            unsafe_allow_html=True,
        )
    with c3:
        st.markdown(
            f'<div style="background:rgba(255,165,0,0.2); border-left:4px solid #FFA500; border-radius:8px; padding:12px;">'
            f'<span style="font-size:0.85rem; color:#aaa;">🟠 Tiède</span><br>'
            f'<span style="font-size:1.8rem; font-weight:700; color:#FFA500;">{n_tiede}</span></div>',
            unsafe_allow_html=True,
        )
    with c4:
        st.markdown(
            f'<div style="background:rgba(255,220,50,0.15); border-left:4px solid #FFDC32; border-radius:8px; padding:12px;">'
            f'<span style="font-size:0.85rem; color:#aaa;">🟡 À surveiller</span><br>'
            f'<span style="font-size:1.8rem; font-weight:700; color:#FFDC32;">{n_surveiller}</span></div>',
            unsafe_allow_html=True,
        )
    with c5:
        st.metric("🖱️ Clics", f"{int(result['Clics'].sum()):,}")
    with c6:
        st.metric("👁️ Impressions", f"{int(result['Impressions'].sum()):,}")

    st.divider()

    # ── Sous-onglets ──
    stab1, stab2, stab3, stab4 = st.tabs([
        "🔥 Points Chauds",
        "📋 Tableau complet",
        "📊 Statistiques",
        "🔍 Recherche",
    ])

    with stab1:
        st.subheader("🔥 Pages à optimiser en priorité")
        hot = result[result["Statut"].isin(["🔥 Chaud", "🟠 Tiède"])]
        if len(hot) > 0:
            st.dataframe(
                hot[display_cols],
                use_container_width=True,
                height=500,
                column_config=COLUMN_CONFIG,
                hide_index=True,
            )
        else:
            st.success("🎉 Aucun point chaud détecté !")

    with stab2:
        st.subheader("📋 Tableau complet")
        statut_filter = st.multiselect(
            "Filtrer par statut",
            options=result["Statut"].unique().tolist(),
            default=result["Statut"].unique().tolist(),
            key=f"statut_filter_{tab_key_suffix}",
        )
        filtered = result[result["Statut"].isin(statut_filter)]
        st.dataframe(
            filtered[display_cols],
            use_container_width=True,
            height=600,
            column_config=COLUMN_CONFIG,
            hide_index=True,
        )

    with stab3:
        st.subheader("📊 Répartition des statuts")
        statut_counts = result["Statut"].value_counts()
        st.bar_chart(statut_counts)

        st.subheader("📊 Distribution des CTR")
        if "CTR" in result.columns:
            bins = pd.cut(
                result["CTR"],
                bins=[0, 1, 2, 3, 5, 10, 100],
                labels=["0-1%", "1-2%", "2-3%", "3-5%", "5-10%", "10%+"],
                include_lowest=True,
            )
            st.bar_chart(bins.value_counts().sort_index())

        st.subheader(f"📊 Top 10 pages par impressions (CTR < 5%)")
        low_ctr = result[result["CTR"] < 5].nlargest(10, "Impressions")
        if len(low_ctr) > 0:
            top_cols = ["URL", info["tag_col"], "Meilleurs mot-clés GSC", "Impressions", "CTR"]
            top_cols = [c for c in top_cols if c in result.columns]
            st.dataframe(
                low_ctr[top_cols],
                use_container_width=True,
                column_config=COLUMN_CONFIG,
                hide_index=True,
            )

    with stab4:
        st.subheader("🔍 Recherche par URL ou mot-clé")
        search = st.text_input("Rechercher", placeholder="URL, mot-clé, balise…", key=f"search_{tab_key_suffix}")
        if search:
            mask = pd.Series(False, index=result.index)
            for col in search_cols:
                if col in result.columns:
                    mask = mask | result[col].astype(str).str.contains(
                        search, case=False, na=False
                    )
            found = result[mask]
            st.write(f"**{len(found)}** résultat(s)")
            st.dataframe(
                found[display_cols],
                use_container_width=True,
                height=400,
                column_config=COLUMN_CONFIG,
                hide_index=True,
            )


def main():
    # ── Header ──
    st.title("🔥 Hotspot Mapper — Mapping des Points Chauds")
    st.caption("Identifie automatiquement les pages à fort potentiel d'optimisation Title, H1, Meta Description / CTR.")

    # ── Sidebar : Upload ──
    with st.sidebar:
        st.header("📁 Import des fichiers")

        st.subheader("1️⃣ Screaming Frog — internal_html")
        st.caption("Export CSV connecté à l'API GSC")
        file_sf = st.file_uploader(
            "Fichier internal_html (.csv)",
            type=["csv"],
            key="sf_upload",
            help="Export > Internal > HTML depuis Screaming Frog, connecté à l'API GSC",
        )

        st.divider()

        st.subheader("2️⃣ Consolidation données GSC")
        st.caption("Fichier avec tous les mots-clés par URL")
        file_gsc = st.file_uploader(
            "Fichier consolidation GSC (.xlsx / .csv)",
            type=["xlsx", "csv"],
            key="gsc_upload",
            help="Fichier consolidé contenant les mots-clés GSC par URL",
        )

        st.divider()

        # ── Filtres ──
        st.header("🎛️ Filtres")

        filter_lang = st.multiselect(
            "Filtrer par langue (segment URL)",
            options=["/fr/", "/en/", "/es/", "/de/", "/it/"],
            default=[],
            help="Filtre les URLs contenant ces segments",
        )

        filter_indexable = st.checkbox("Pages indexables uniquement", value=True)
        filter_http_200 = st.checkbox("Code HTTP 200 uniquement", value=True)

        filter_min_impressions = st.number_input(
            "Impressions minimum", min_value=0, value=0, step=10,
        )

        filter_min_clicks = st.number_input(
            "Clics minimum", min_value=0, value=0, step=5,
        )

        st.divider()

        # ── Filtre mots-clés de marque ──
        st.header("🏷️ Mots-clés marque")
        brand_regex = st.text_input(
            "Regex mots-clés à exclure",
            value="",
            placeholder="ranch.?computing|ranchcomputing",
            help=(
                "Regex (insensible à la casse) pour supprimer les mots-clés de marque "
                "de la colonne 'Meilleurs mot-clés GSC'. "
                "Exemples : `ranch.?computing|ranchcomputing`, `marque1|marque2`"
            ),
            key="brand_regex",
        )
        if brand_regex:
            try:
                re.compile(brand_regex)
                st.caption(f"✅ Regex valide : `{brand_regex}`")
            except re.error as e:
                st.error(f"❌ Regex invalide : {e}")
                brand_regex = ""

        st.divider()

        # ── Google Sheets ──
        st.header("📊 Google Sheets")
        has_credentials = _get_credentials_path() is not None
        if not has_credentials:
            st.warning("⚠️ credentials.json introuvable")
        else:
            st.caption("☁️ Mise à jour de votre Google Sheet personnel")

    # ── Accueil ──
    if file_sf is None and file_gsc is None:
        st.info("👈 Commencez par uploader vos fichiers dans la barre latérale.")
        with st.expander("📖 Mode d'emploi", expanded=True):
            st.markdown("""
### Comment utiliser cet outil ?

---

#### 📁 Étape 1 — Uploader les fichiers (sidebar gauche)

**Fichier 1 : `internal_html.csv`** (Screaming Frog)
- Export CSV du crawl connecté à l'API GSC
- Fournit : URLs, Title, H1, Meta Description, longueurs, Code HTTP, Indexabilité

**Fichier 2 : Consolidation GSC** (`.xlsx` ou `.csv`)
- Fichier consolidé avec tous les mots-clés GSC par URL
- Fournit : Clics, Impressions, Mots-clés par page

> ⚠️ Les **deux fichiers** sont nécessaires pour lancer la fusion.

---

#### 🎛️ Étape 2 — Configurer les filtres (sidebar gauche)

| Filtre | Description | Recommandé |
|--------|-------------|------------|
| **Langue** | Segment URL : `/fr/`, `/en/`… (sites multilingues) | Selon le site |
| **Indexables uniquement** | Ne garde que les pages indexables | ✅ Oui |
| **HTTP 200 uniquement** | Exclut redirections et erreurs | ✅ Oui |
| **Impressions min.** | Seuil minimum d'impressions | 0 à 100 selon taille |
| **Clics min.** | Seuil minimum de clics | 0 à 50 selon taille |
| **Regex marque** | Exclut les mots-clés de marque (`marque1\\|marque2`) | À remplir |

> ⚠️ **Filtre langue** : le filtre cherche le segment **exact** dans l'URL (ex: `/fr/`). Vérifiez que vos URLs contiennent bien ce pattern. Si votre site utilise un autre format (ex: `/fr-fr/`, `?lang=fr`, sous-domaine `fr.site.com`), ce filtre ne fonctionnera pas — laissez-le vide.

---

#### 🔥 Étape 3 — Lire les résultats

L'outil fusionne automatiquement les données et calcule un **Score de Chaleur** :

`Score = (Impressions / Impressions max du site) × 100 × (1 − CTR)`

**Pourquoi ce score ?** Une page avec beaucoup d'impressions mais un CTR très bas = fort potentiel inexploité. Le score normalise les impressions (rapportées au max du site) et les pondère par l'inverse du CTR. Plus le score est élevé, plus la page a de potentiel à gagner en optimisant ses balises.

**Qu'est-ce que ça dit de votre contenu ?**
- **Score élevé (🔥 Chaud)** : Google juge votre page pertinente pour ces mots-clés (il l'affiche souvent), mais **les internautes ne cliquent pas**. Le problème vient de ce que l'utilisateur voit dans les résultats de recherche : votre **Title et/ou Meta Description ne donnent pas envie de cliquer**. Ils ne correspondent pas à l'intention de recherche, sont trop génériques, ou ne mettent pas en avant la valeur de la page.
- **Score moyen (🟠 Tiède)** : même diagnostic, à moindre échelle. La page a de la visibilité mais sous-performe en clics — une optimisation des balises peut rapidement augmenter le trafic.
- **Score faible (✅ OK)** : soit la page a peu d'impressions (faible visibilité), soit elle a un bon CTR (les balises font leur travail). Pas d'action prioritaire.

**En résumé** : le score identifie les pages où **le contenant (balises) ne rend pas justice au contenu**. Google vous donne de la visibilité, mais vos balises ne convertissent pas cette visibilité en clics. Optimiser ces balises = gains de trafic rapides sans toucher au contenu de la page.

| Statut | Condition | Action |
|--------|-----------|--------|
| 🔥 **Chaud** | ≥500 impressions, CTR < 2% | Optimiser en urgence — fort potentiel de gain |
| 🟠 **Tiède** | ≥100 impressions, CTR < 3% | À optimiser — potentiel significatif |
| 🟡 **À surveiller** | ≥20 impressions, CTR < 5% | Surveiller — potentiel modéré |
| ✅ **OK** | Reste | Pas d'action prioritaire |

**3 onglets principaux** : MDPC Title / MDPC H1 / MDPC Meta Description

**Dans chaque onglet — 4 sous-onglets** :
- 🔥 **Points Chauds** : pages Chaud + Tiède, triées par score décroissant
- 📋 **Tableau complet** : toutes les pages, filtrable par statut
- 📊 **Statistiques** : répartition des statuts, distribution CTR, Top 10 impressions
- 🔍 **Recherche** : recherche libre par URL, mot-clé ou balise

---

#### 📥 Étape 4 — Exporter

| Export | Contenu |
|--------|---------|
| **Excel complet (3 feuilles)** | 1 fichier .xlsx avec Title + H1 + Meta Desc. *(recommandé)* |
| **Excel Title seul** | 1 fichier .xlsx, onglet Title uniquement |
| **CSV (Title)** | Export CSV rapide |

Le fichier Excel contient des **colonnes vides** prêtes pour la Phase 2 :
`Positions` · `Balise proposée` · `Longueur proposée`

---

#### ☁️ Étape 5 — Export Google Sheets (optionnel)

Si `credentials.json` est présent, le bouton **"Mettre à jour mon Google Sheet"** apparaît.
Il pousse les 3 feuilles MDPC + un onglet Config (paramètres DataForSEO pré-remplis).

> ⚠️ **Pour utiliser un autre Google Sheet** : modifiez la variable `PERSONAL_SHEET_URL` dans `app.py` (ligne ~600) avec l'URL de votre fichier. Le compte de service (`credentials.json`) doit avoir accès en éditeur à ce fichier — partagez le Google Sheet avec l'adresse email du compte de service (visible dans le fichier JSON, champ `client_email`).
            """)
        return

    # ── Parse fichiers ──
    df_sf = None
    df_gsc = None

    if file_sf is not None:
        try:
            df_sf_raw = pd.read_csv(file_sf, encoding="utf-8", on_bad_lines="skip")
            df_sf = parse_screaming_frog(df_sf_raw)
            st.sidebar.success(f"✅ SF : {len(df_sf)} URLs")
        except Exception as e:
            st.sidebar.error(f"❌ Erreur SF : {e}")

    if file_gsc is not None:
        try:
            if file_gsc.name.endswith(".xlsx"):
                df_gsc_raw = pd.read_excel(file_gsc)
            else:
                df_gsc_raw = pd.read_csv(file_gsc, encoding="utf-8", on_bad_lines="skip")
            df_gsc = parse_gsc_consolidation(df_gsc_raw)
            st.sidebar.success(f"✅ GSC : {len(df_gsc)} URLs")
        except Exception as e:
            st.sidebar.error(f"❌ Erreur GSC : {e}")

    # ── Aperçu individuel ──
    if df_sf is not None and df_gsc is None:
        st.subheader("📊 Aperçu — Screaming Frog")
        st.info("⬆️ Uploadez aussi le fichier de consolidation GSC pour la fusion complète.")
        st.dataframe(df_sf, use_container_width=True, height=500)
        return

    if df_gsc is not None and df_sf is None:
        st.subheader("📊 Aperçu — Consolidation GSC")
        st.info("⬆️ Uploadez aussi le fichier Screaming Frog pour la fusion complète.")
        st.dataframe(df_gsc, use_container_width=True, height=500)
        return

    # ── FUSION — 3 FEUILLES ──
    if df_sf is not None and df_gsc is not None:

        # Générer les 3 DataFrames
        result_title = merge_and_build(df_sf, df_gsc, sheet_type="title", brand_regex=brand_regex)
        result_h1 = merge_and_build(df_sf, df_gsc, sheet_type="h1", brand_regex=brand_regex)
        result_meta = merge_and_build(df_sf, df_gsc, sheet_type="meta_description", brand_regex=brand_regex)

        # Appliquer les filtres
        results = {}
        for sheet_type, raw_df in [("title", result_title), ("h1", result_h1), ("meta_description", result_meta)]:
            df_filtered = _apply_filters(
                raw_df, filter_lang, filter_indexable, filter_http_200,
                filter_min_impressions, filter_min_clicks,
            )
            df_filtered = detect_hotspots(df_filtered)
            df_filtered = df_filtered.sort_values("Score Chaleur", ascending=False).reset_index(drop=True)
            results[sheet_type] = df_filtered

        # ── Vue d'ensemble rapide ──
        st.subheader("📈 Vue d'ensemble")
        ov1, ov2, ov3 = st.columns(3)
        for col_widget, (stype, sdf) in zip([ov1, ov2, ov3], results.items()):
            info = SHEET_TYPES[stype]
            hot_count = len(sdf[sdf["Statut"].isin(["🔥 Chaud", "🟠 Tiède"])])
            with col_widget:
                st.metric(info["label"], f"{hot_count} points chauds", delta=f"{len(sdf)} URLs")

        st.divider()

        # ── Onglets principaux : 1 par type de feuille ──
        main_tab_title, main_tab_h1, main_tab_meta = st.tabs([
            "🏷️ MDPC Title",
            "📝 MDPC H1",
            "📋 MDPC Meta Description",
        ])

        with main_tab_title:
            _render_sheet_tab(results["title"], "title", tab_key_suffix="title")

        with main_tab_h1:
            _render_sheet_tab(results["h1"], "h1", tab_key_suffix="h1")

        with main_tab_meta:
            _render_sheet_tab(results["meta_description"], "meta_description", tab_key_suffix="meta")

        # ── Export ──
        st.divider()
        st.subheader("📥 Export")

        ts = datetime.now().strftime("%Y-%m-%d_%H-%M")

        # Préparer les DataFrames d'export pour chaque feuille
        export_dfs = {}
        for stype, sdf in results.items():
            ecols = [c for c in _get_export_cols(stype) if c in sdf.columns]
            export_dfs[stype] = sdf[ecols]

        col_dl1, col_dl2, col_dl3 = st.columns(3)

        with col_dl1:
            # Export Excel multi-feuilles (3 onglets dans 1 fichier)
            excel_title = _prepare_excel_df(export_dfs["title"])
            excel_h1 = _prepare_excel_df(export_dfs["h1"])
            excel_meta = _prepare_excel_df(export_dfs["meta_description"])

            st.download_button(
                label="📥 Excel complet (3 feuilles)",
                data=to_excel_multi_sheets(excel_title, excel_h1, excel_meta),
                file_name=f"MDPC_mapping_complet_{ts}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary",
                use_container_width=True,
            )

        with col_dl2:
            # Export Excel Title seul
            excel_title_only = _prepare_excel_df(export_dfs["title"])
            st.download_button(
                label="📥 Excel Title seul",
                data=to_excel_download(excel_title_only),
                file_name=f"MDPC_title_{ts}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )

        with col_dl3:
            # Export CSV (Title par défaut)
            st.download_button(
                label="📥 CSV (Title)",
                data=export_dfs["title"].to_csv(index=False).encode("utf-8"),
                file_name=f"MDPC_title_{ts}.csv",
                mime="text/csv",
                use_container_width=True,
            )

        # ── Export Google Sheets ──
        if has_credentials:
            st.divider()

            if st.button(
                "☁️ Mettre à jour mon Google Sheet",
                type="primary",
                use_container_width=True,
                help="Met à jour votre Google Sheet personnel avec les données MDPC + Config",
            ):
                with st.spinner("⏳ Mise à jour du Google Sheet en cours..."):
                    try:
                        gs_title = export_dfs["title"].copy()
                        gs_h1 = export_dfs["h1"].copy()
                        gs_meta = export_dfs["meta_description"].copy()

                        url = _update_personal_sheet(gs_title, gs_h1, gs_meta)
                        st.success(f"✅ Google Sheet mis à jour ! [Ouvrir le fichier]({url})")
                    except Exception as e:
                        st.error(f"❌ Erreur : {e}")


if __name__ == "__main__":
    main()
