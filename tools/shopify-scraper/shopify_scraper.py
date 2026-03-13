#!/usr/bin/env python3
"""
╔══════════════════════════════════════════════════════════╗
║       SHOPIFY PRODUCT SCRAPER — Analyse complète         ║
║                                                          ║
║  Usage:                                                  ║
║    python3 shopify_scraper.py balibaris.com               ║
║    python3 shopify_scraper.py www.example-shop.com        ║
║                                                          ║
║  Ou double-clic sur le fichier .command (Mac)            ║
╚══════════════════════════════════════════════════════════╝
"""

import json
import re
import ssl
import sys
import shutil
import urllib.request
import urllib.error
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path

# Fix SSL pour macOS
SSL_CTX = ssl.create_default_context()
SSL_CTX.check_hostname = False
SSL_CTX.verify_mode = ssl.CERT_NONE

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import BarChart, PieChart, Reference
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


# ─────────────────────────────────────────────
# CONFIGURATION
# ─────────────────────────────────────────────
DEFAULT_STORE = "www.balibaris.com"
PRODUCTS_PER_PAGE = 250
DRIVE_SUBFOLDER = "Scraping"


# ─────────────────────────────────────────────
# GOOGLE DRIVE
# ─────────────────────────────────────────────
def find_google_drive_path():
    """Détecte automatiquement le dossier Google Drive sur Mac/Windows/Linux."""
    home = Path.home()
    cloud_storage = home / "Library" / "CloudStorage"
    if cloud_storage.exists():
        matches = list(cloud_storage.glob("GoogleDrive-*"))
        if matches:
            for drive_root in matches:
                for drive_name in ["Mon Drive", "My Drive", "Mi unidad"]:
                    candidate = drive_root / drive_name
                    if candidate.exists():
                        return candidate
            return matches[0]
    win_path = home / "Google Drive"
    if win_path.exists():
        return win_path
    old_mac = Path("/Volumes/GoogleDrive/Mon Drive")
    if old_mac.exists():
        return old_mac
    return None


def copy_to_google_drive(local_path, filename):
    """Copie le fichier vers Google Drive/Scraping/ si disponible."""
    drive_root = find_google_drive_path()
    if not drive_root:
        print("\n  ⚠️  Google Drive non détecté — fichier sauvegardé localement uniquement")
        print("     → Installe Google Drive for Desktop pour la sync automatique")
        return None
    drive_folder = drive_root / DRIVE_SUBFOLDER
    drive_folder.mkdir(parents=True, exist_ok=True)
    dest = drive_folder / filename
    shutil.copy2(local_path, dest)
    print(f"\n  ☁️  Copié sur Google Drive : {dest}")
    return dest


# ─────────────────────────────────────────────
# 1. RÉCUPÉRATION DES PRODUITS
# ─────────────────────────────────────────────
def fetch_all_products(store_domain):
    """Récupère tous les produits via l'API Shopify /products.json"""
    all_products = []
    page = 1
    store_domain = store_domain.replace("https://", "").replace("http://", "").strip("/")
    if not store_domain.startswith("www."):
        store_domain = "www." + store_domain
    base_url = f"https://{store_domain}/products.json?limit={PRODUCTS_PER_PAGE}"

    print(f"\n🔍 Scraping de {store_domain}...")
    print("─" * 50)

    while True:
        url = f"{base_url}&page={page}"
        try:
            req = urllib.request.Request(url, headers={
                "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7)"
            })
            with urllib.request.urlopen(req, timeout=30, context=SSL_CTX) as response:
                data = json.loads(response.read().decode())
            products = data.get("products", [])
            if not products:
                break
            all_products.extend(products)
            print(f"   📦 Page {page} : {len(products)} produits récupérés (total: {len(all_products)})")
            page += 1
        except urllib.error.HTTPError as e:
            if e.code == 429:
                print("   ⏳ Rate limit atteint, pause de 2s...")
                import time; time.sleep(2)
                continue
            else:
                print(f"   ❌ Erreur HTTP {e.code}: {e.reason}")
                break
        except Exception as e:
            print(f"   ❌ Erreur: {e}")
            break

    print(f"\n✅ Total: {len(all_products)} produits récupérés")
    return all_products


# ─────────────────────────────────────────────
# 2. EXTRACTION DES MATIÈRES
# ─────────────────────────────────────────────
def extract_materials_from_tags(tags):
    materials = []
    for tag in tags:
        tag_lower = tag.lower().strip()
        if tag_lower.startswith("matiere:") or tag_lower.startswith("matière:"):
            mat = tag_lower.split(":", 1)[1].strip().capitalize()
            materials.append(mat)
    return materials


def extract_materials_from_description(html_body):
    if not html_body:
        return []
    text = re.sub(r'<[^>]+>', ' ', html_body).replace('&nbsp;', ' ').replace('\n', ' ')
    materials = []
    patterns = [
        r'(\d+)%\s*([\wéèêë]+)',
        r'(coton|lin|laine|soie|cachemire|polyester|viscose|elasthanne|nylon|modal|lyocell|tencel)',
    ]
    for pattern in patterns:
        matches = re.findall(pattern, text.lower())
        for match in matches:
            mat = (match[1] if isinstance(match, tuple) else match).strip().capitalize()
            if mat and mat not in materials:
                materials.append(mat)
    return materials


def extract_composition_detail(html_body):
    if not html_body:
        return ""
    text = re.sub(r'<[^>]+>', ' ', html_body)
    matches = re.findall(r'\d+%\s*[\wéèêë]+', text.lower())
    return ", ".join(matches) if matches else ""


# ─────────────────────────────────────────────
# 3. ANALYSE DES DONNÉES
# ─────────────────────────────────────────────
def analyze_products(products):
    results = {
        "products_detail": [],
        "materials_count": Counter(),
        "type_count": Counter(),
        "color_count": Counter(),
        "collection_count": Counter(),
        "coupe_count": Counter(),
        "forme_count": Counter(),
        "material_by_type": defaultdict(Counter),
        "coupe_by_type": defaultdict(Counter),
        "price_by_material": defaultdict(list),
        "taxonomy": defaultdict(set),
        "combos_type_mat": Counter(),
        "combos_type_col": Counter(),
        "combos_type_coll": Counter(),
        "combos_type_mat_col": Counter(),
        "combos_type_coupe": Counter(),
        "total": len(products),
    }

    for p in products:
        tags = p.get("tags", [])
        title = p.get("title", "")
        product_type = p.get("product_type", "Non défini")
        body = p.get("body_html", "")

        # Matières
        materials = extract_materials_from_tags(tags)
        if not materials:
            materials = extract_materials_from_description(body)

        # Couleur
        colors = [t.split(":", 1)[1].strip().capitalize()
                  for t in tags if t.lower().startswith("couleur:")]

        # Collections
        collections = [t.split(":", 1)[1].strip()
                       for t in tags if t.lower().startswith("collection:")]

        # Coupes
        coupes = [t.split(":", 1)[1].strip().capitalize()
                  for t in tags if t.lower().startswith("coupe:")]

        # Formes
        formes = [t.split(":", 1)[1].strip().capitalize()
                  for t in tags if t.lower().startswith("forme:")]

        # Prix
        variants = p.get("variants", [])
        price = None
        compare_price = None
        if variants:
            price = float(variants[0].get("price", 0))
            cp = variants[0].get("compare_at_price")
            if cp:
                compare_price = float(cp)

        composition = extract_composition_detail(body)

        results["products_detail"].append({
            "title": title, "type": product_type,
            "materials": ", ".join(materials) if materials else "Non renseigné",
            "composition": composition,
            "colors": ", ".join(colors) if colors else "Non renseigné",
            "coupes": ", ".join(coupes) if coupes else "",
            "formes": ", ".join(formes) if formes else "",
            "price": price, "compare_price": compare_price,
            "url": f"https://{DEFAULT_STORE}/products/{p.get('handle', '')}",
            "tags": ", ".join(tags),
        })

        # Compteurs
        results["type_count"][product_type] += 1
        for mat in materials:
            results["materials_count"][mat] += 1
            results["material_by_type"][mat][product_type] += 1
            if price:
                results["price_by_material"][mat].append(price)
        for col in colors:
            results["color_count"][col] += 1
        for coll in collections:
            results["collection_count"][coll] += 1
        for coupe in coupes:
            results["coupe_count"][coupe] += 1
            results["coupe_by_type"][coupe][product_type] += 1
        for forme in formes:
            results["forme_count"][forme] += 1

        # Combinaisons réelles
        if product_type:
            for mat in materials:
                results["combos_type_mat"][f"{product_type.lower()} {mat.lower()}"] += 1
            for col in colors:
                results["combos_type_col"][f"{product_type.lower()} {col.lower()}"] += 1
            for coll in collections:
                results["combos_type_coll"][f"{product_type.lower()} {coll.lower()}"] += 1
            for coupe in coupes:
                results["combos_type_coupe"][f"{product_type.lower()} {coupe.lower()}"] += 1
            for mat in materials:
                for col in colors:
                    results["combos_type_mat_col"][f"{product_type.lower()} {mat.lower()} {col.lower()}"] += 1

        # Taxonomie
        results["taxonomy"]["Type de produit"].add(product_type)
        for tag in tags:
            tag_str = tag if isinstance(tag, str) else str(tag)
            if ":" in tag_str:
                prefix, value = tag_str.split(":", 1)
                prefix = prefix.strip().lower()
                value = value.strip().capitalize()
                prefix_map = {
                    "matiere": "Matières", "matière": "Matières",
                    "couleur": "Couleurs", "couleurs": "Couleurs", "coloris": "Couleurs",
                    "collection": "Collections", "main_collection": "Catégories principales",
                    "coupe": "Coupes", "forme": "Formes", "motif": "Motifs",
                    "gender": "Genre", "style": "Styles",
                    "size_guide": "Guides de tailles", "saison": "Saisons",
                }
                col_name = prefix_map.get(prefix)
                if col_name:
                    results["taxonomy"][col_name].add(value)

    return results


# ─────────────────────────────────────────────
# 4. EXPORT EXCEL
# ─────────────────────────────────────────────
def style_header(ws, row, max_col):
    header_fill = PatternFill(start_color="1B2A4A", end_color="1B2A4A", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin_border = Border(bottom=Side(style="thin", color="1B2A4A"))
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border


def export_to_excel(results, store_name):
    if not HAS_OPENPYXL:
        print("❌ openpyxl non installé. Lance: pip3 install openpyxl")
        return None

    wb = Workbook()
    total = results["total"]

    # ── Onglet 1 : Par Matière ──
    ws1 = wb.active
    ws1.title = "📊 Par Matière"
    headers = ["Matière", "Nb produits", "% du catalogue", "Prix moyen (€)", "Prix min (€)", "Prix max (€)"]
    ws1.append(headers)
    style_header(ws1, 1, len(headers))
    for mat, count in results["materials_count"].most_common():
        prices = results["price_by_material"].get(mat, [])
        avg_p = round(sum(prices) / len(prices), 2) if prices else ""
        min_p = round(min(prices), 2) if prices else ""
        max_p = round(max(prices), 2) if prices else ""
        pct = round(count / total * 100, 1) if total else 0
        ws1.append([mat, count, f"{pct}%", avg_p, min_p, max_p])
    for c in range(1, 7):
        ws1.column_dimensions[get_column_letter(c)].width = 18
    if results["materials_count"]:
        pie = PieChart()
        pie.title = "Répartition par matière"
        pie.style = 10
        data_ref = Reference(ws1, min_col=2, min_row=1, max_row=len(results["materials_count"]) + 1)
        cats_ref = Reference(ws1, min_col=1, min_row=2, max_row=len(results["materials_count"]) + 1)
        pie.add_data(data_ref, titles_from_data=True)
        pie.set_categories(cats_ref)
        pie.width = 18; pie.height = 12
        ws1.add_chart(pie, "H2")

    # ── Onglet 2 : Par Type ──
    ws2 = wb.create_sheet("📦 Par Type")
    ws2.append(["Type de produit", "Nombre", "% du catalogue"])
    style_header(ws2, 1, 3)
    for ptype, count in results["type_count"].most_common():
        pct = round(count / total * 100, 1) if total else 0
        ws2.append([ptype, count, f"{pct}%"])
    ws2.column_dimensions["A"].width = 25
    ws2.column_dimensions["B"].width = 15
    ws2.column_dimensions["C"].width = 15

    # ── Onglet 3 : Matière x Type ──
    ws3 = wb.create_sheet("🔀 Matière x Type")
    all_types = sorted(set(t for counts in results["material_by_type"].values() for t in counts))
    headers3 = ["Matière"] + all_types
    ws3.append(headers3)
    style_header(ws3, 1, len(headers3))
    for mat, type_counts in sorted(results["material_by_type"].items()):
        ws3.append([mat] + [type_counts.get(t, 0) for t in all_types])
    for c in range(1, len(headers3) + 1):
        ws3.column_dimensions[get_column_letter(c)].width = 16

    # ── Onglet 4 : Par Couleur ──
    ws4 = wb.create_sheet("🎨 Par Couleur")
    ws4.append(["Couleur", "Nombre", "% du catalogue"])
    style_header(ws4, 1, 3)
    for color, count in results["color_count"].most_common():
        pct = round(count / total * 100, 1) if total else 0
        ws4.append([color, count, f"{pct}%"])
    ws4.column_dimensions["A"].width = 20
    ws4.column_dimensions["B"].width = 15
    ws4.column_dimensions["C"].width = 15

    # ── Onglet 5 : Par Coupe ──
    ws5 = wb.create_sheet("📐 Par Coupe")
    # 5a : Comptage simple
    ws5.append(["Coupe", "Nb produits", "% du catalogue"])
    style_header(ws5, 1, 3)
    for coupe, count in results["coupe_count"].most_common():
        pct = round(count / total * 100, 1) if total else 0
        ws5.append([coupe, count, f"{pct}%"])
    ws5.column_dimensions["A"].width = 22
    ws5.column_dimensions["B"].width = 15
    ws5.column_dimensions["C"].width = 15

    # 5b : Camembert coupes
    if results["coupe_count"]:
        pie_c = PieChart()
        pie_c.title = "Répartition par coupe"
        pie_c.style = 10
        n_coupes = len(results["coupe_count"])
        data_ref = Reference(ws5, min_col=2, min_row=1, max_row=n_coupes + 1)
        cats_ref = Reference(ws5, min_col=1, min_row=2, max_row=n_coupes + 1)
        pie_c.add_data(data_ref, titles_from_data=True)
        pie_c.set_categories(cats_ref)
        pie_c.width = 16; pie_c.height = 11
        ws5.add_chart(pie_c, "E2")

    # 5c : Croisement Coupe x Type (en dessous)
    start_row = len(results["coupe_count"]) + 4
    ws5.cell(row=start_row, column=1, value="CROISEMENT COUPE × TYPE DE PRODUIT").font = Font(bold=True, size=12, color="1B2A4A")
    start_row += 1
    all_types_coupe = sorted(set(t for counts in results["coupe_by_type"].values() for t in counts))
    headers5c = ["Coupe"] + all_types_coupe
    for ci, h in enumerate(headers5c, 1):
        ws5.cell(row=start_row, column=ci, value=h)
    style_header(ws5, start_row, len(headers5c))
    start_row += 1
    for coupe, type_counts in sorted(results["coupe_by_type"].items()):
        row_data = [coupe] + [type_counts.get(t, 0) for t in all_types_coupe]
        for ci, val in enumerate(row_data, 1):
            ws5.cell(row=start_row, column=ci, value=val)
        start_row += 1

    # ── Onglet 6 : Tous les produits ──
    ws6 = wb.create_sheet("📋 Tous les produits")
    headers6 = ["Produit", "Type", "Matières", "Composition", "Couleurs", "Coupes", "Formes", "Prix (€)", "Ancien prix (€)", "URL"]
    ws6.append(headers6)
    style_header(ws6, 1, len(headers6))
    for p in results["products_detail"]:
        ws6.append([
            p["title"], p["type"], p["materials"], p["composition"],
            p["colors"], p["coupes"], p["formes"],
            p["price"], p["compare_price"], p["url"]
        ])
    widths6 = [40, 18, 25, 30, 18, 18, 18, 12, 15, 50]
    for ci, w in enumerate(widths6, 1):
        ws6.column_dimensions[get_column_letter(ci)].width = w

    # ── Onglet 7 : Taxonomie ──
    ws7 = wb.create_sheet("🗂 Taxonomie")
    taxonomy = results.get("taxonomy", {})
    col_order = [
        "Type de produit", "Catégories principales", "Matières", "Couleurs",
        "Coupes", "Formes", "Motifs", "Collections", "Saisons", "Genre",
        "Guides de tailles", "Styles",
    ]
    active_cols = [c for c in col_order if c in taxonomy and taxonomy[c]]
    for c in sorted(taxonomy.keys()):
        if c not in active_cols and taxonomy[c]:
            active_cols.append(c)

    if active_cols:
        ws7.append(active_cols)
        style_header(ws7, 1, len(active_cols))
        sorted_cols = {col: sorted(taxonomy[col]) for col in active_cols}
        max_rows = max(len(v) for v in sorted_cols.values())
        for row_idx in range(max_rows):
            row_data = []
            for col in active_cols:
                vals = sorted_cols[col]
                row_data.append(vals[row_idx] if row_idx < len(vals) else "")
            ws7.append(row_data)
        for ci in range(1, len(active_cols) + 1):
            ws7.column_dimensions[get_column_letter(ci)].width = 25
        ws7.append([])
        count_row = [f"({len(sorted_cols[col])} valeurs)" for col in active_cols]
        ws7.append(count_row)
        for ci in range(1, len(active_cols) + 1):
            ws7.cell(row=max_rows + 3, column=ci).font = Font(italic=True, color="888888")

    # ── Onglet 8 : Mots-clés SEO ──
    ws8 = wb.create_sheet("🔑 Mots-clés SEO")
    product_types = sorted(taxonomy.get("Type de produit", set()))
    cross_attributes = [
        ("Catégories principales", "cat + sous cat"),
        ("Matières", "cat + matières"),
        ("Couleurs", "cat + couleurs"),
        ("Coupes", "cat + coupes"),
        ("Formes", "cat + formes"),
        ("Motifs", "cat + motifs"),
        ("Styles", "cat + styles"),
        ("Saisons", "cat + saisons"),
    ]
    if product_types:
        headers8 = ["cat"]
        active_cross = []
        for attr_key, col_name in cross_attributes:
            if attr_key in taxonomy and taxonomy[attr_key]:
                headers8.append(col_name)
                active_cross.append((attr_key, col_name))
        ws8.append(headers8)
        style_header(ws8, 1, len(headers8))
        all_combos = {"cat": product_types}
        for attr_key, col_name in active_cross:
            combos = []
            for ptype in product_types:
                for val in sorted(taxonomy[attr_key]):
                    combos.append(f"{ptype.lower()} {val.lower()}")
            all_combos[col_name] = combos
        max_kw_rows = max(len(v) for v in all_combos.values())
        for row_idx in range(max_kw_rows):
            row_data = [all_combos[h][row_idx] if row_idx < len(all_combos[h]) else "" for h in headers8]
            ws8.append(row_data)
        for ci in range(1, len(headers8) + 1):
            ws8.column_dimensions[get_column_letter(ci)].width = 30
        ws8.append([])
        ws8.append([f"({len(all_combos[h])} mots-clés)" for h in headers8])
        for ci in range(1, len(headers8) + 1):
            ws8.cell(row=max_kw_rows + 3, column=ci).font = Font(italic=True, color="888888")

    # ── Onglet 9 : Top Combinaisons ──
    ws9 = wb.create_sheet("🏆 Top Combinaisons")
    combo_sections = [
        ("combos_type_mat", "🧵 Type + Matière", "Ex: chemise coton, pull laine..."),
        ("combos_type_col", "🎨 Type + Couleur", "Ex: chemise bleu, pantalon beige..."),
        ("combos_type_coupe", "📐 Type + Coupe", "Ex: chemise slim, pantalon ajustée..."),
        ("combos_type_coll", "📁 Type + Collection", "Ex: chemise casual, pull mailles..."),
        ("combos_type_mat_col", "🔗 Type + Matière + Couleur", "Ex: chemise coton bleu..."),
    ]
    current_row = 1
    for combo_key, section_title, description in combo_sections:
        combo_data = results.get(combo_key, Counter())
        if not combo_data:
            continue
        ws9.cell(row=current_row, column=1, value=section_title).font = Font(bold=True, size=13, color="1B2A4A")
        ws9.cell(row=current_row, column=3, value=description).font = Font(italic=True, color="888888")
        current_row += 1
        ws9.cell(row=current_row, column=1, value="Combinaison")
        ws9.cell(row=current_row, column=2, value="Nb produits")
        ws9.cell(row=current_row, column=3, value="% du total")
        style_header(ws9, current_row, 3)
        current_row += 1
        top1_count = combo_data.most_common(1)[0][1] if combo_data else 1
        for combo, count in combo_data.most_common(50):
            ws9.cell(row=current_row, column=1, value=combo)
            ws9.cell(row=current_row, column=2, value=count)
            pct = round(count / total * 100, 1) if total else 0
            ws9.cell(row=current_row, column=3, value=f"{pct}%")
            bar = "█" * min(int(count / max(1, top1_count) * 20), 20)
            ws9.cell(row=current_row, column=4, value=bar).font = Font(color="4472C4")
            current_row += 1
        ws9.cell(row=current_row, column=1, value=f"Total combinaisons uniques : {len(combo_data)}").font = Font(italic=True, color="888888")
        current_row += 2
    ws9.column_dimensions["A"].width = 40
    ws9.column_dimensions["B"].width = 14
    ws9.column_dimensions["C"].width = 12
    ws9.column_dimensions["D"].width = 25

    # ── Sauvegarde + Google Drive ──
    timestamp = datetime.now().strftime("%Y%m%d_%H%M")
    filename = f"analyse_{store_name}_{timestamp}.xlsx"
    output_path = Path(__file__).parent / filename
    wb.save(output_path)
    drive_path = copy_to_google_drive(output_path, filename)
    return output_path, drive_path


# ─────────────────────────────────────────────
# 5. AFFICHAGE CONSOLE
# ─────────────────────────────────────────────
def print_summary(results):
    print("\n" + "═" * 55)
    print("  📊  RÉSUMÉ DE L'ANALYSE")
    print("═" * 55)
    print(f"\n  Total produits : {results['total']}")

    print(f"\n  🧵 MATIÈRES ({len(results['materials_count'])} différentes) :")
    for mat, count in results["materials_count"].most_common(15):
        pct = round(count / results["total"] * 100, 1)
        bar = "█" * int(pct / 2)
        print(f"     {mat:<20} {count:>4} produits ({pct:>5}%)  {bar}")

    print(f"\n  📦 TYPES DE PRODUITS :")
    for ptype, count in results["type_count"].most_common(10):
        print(f"     {ptype:<25} {count:>4}")

    print(f"\n  📐 COUPES :")
    for coupe, count in results["coupe_count"].most_common(10):
        print(f"     {coupe:<25} {count:>4}")

    print(f"\n  🎨 TOP 10 COULEURS :")
    for color, count in results["color_count"].most_common(10):
        print(f"     {color:<20} {count:>4}")

    print("\n" + "═" * 55)


# ─────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────
def main():
    global DEFAULT_STORE

    if len(sys.argv) > 1:
        store_domain = sys.argv[1]
    else:
        store_domain = DEFAULT_STORE

    store_name = store_domain.replace("www.", "").replace(".com", "").replace(".fr", "").replace(".", "_")
    DEFAULT_STORE = store_domain

    products = fetch_all_products(store_domain)
    if not products:
        print("❌ Aucun produit trouvé. Vérifiez que le site est bien sur Shopify.")
        sys.exit(1)

    results = analyze_products(products)
    print_summary(results)

    result = export_to_excel(results, store_name)
    if result:
        output_path, drive_path = result
        print(f"\n  📁 Fichier Excel généré : {output_path}")
        if drive_path:
            print(f"  ☁️  Synchronisé sur Google Drive : {drive_path}")
        print("     → Ouvre-le pour voir les graphiques et tableaux détaillés\n")


if __name__ == "__main__":
    main()
